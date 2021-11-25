import sys
import os
from PyQt5 import uic
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QPixmap , QIcon
from PyQt5.QtCore import Qt
import rpyc
import socket
from time import sleep
from datetime import datetime, timedelta
import json
import ctypes
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch , mm
from reportlab.lib.colors import white, black
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from openpyxl.utils import get_column_letter
import subprocess

class Login(QDialog):
    ventana_principal = 0
    def __init__(self, parent):
        super(Login,self).__init__(parent)
        uic.loadUi('login2.ui',self)

        self.setModal(True)
        self.conexion = None
        self.host = None
        self.puerto = None
        self.exito = False
        self.actual = None
        self.datos_usuario = 'USER: NINGUNO BRO'
        self.txt_contra.setEchoMode(QLineEdit.Password)
        self.btn_manual.clicked.connect(self.conectar_manual)
        self.btn_iniciar.clicked.connect(self.iniciar)
        self.btn_manual.clicked.connect
        self.inicializar()
        self.txt_usuario.setFocus()
        

    def inicializar(self):
        actual = os.path.abspath(os.getcwd())
        actual = actual.replace('\\' , '/')
        self.actual = actual
        ruta = actual + '/icono_imagen/madenco logo.png'
        foto = QPixmap(ruta)
        self.lb_logo.setPixmap(foto)

        if os.path.isfile(actual + '/manifest.txt'):
            print('encontrado manifest')
            with open(actual + '/manifest.txt' , 'r', encoding='utf-8') as file:
                lines = file.readlines()
                try:
                    n_host = lines[0].split(':')
                    n_host = n_host[1]
                    host = n_host[:len(n_host)-1]

                    n_port = lines[1].split(':')
                    n_port = n_port[1]
                    port = n_port[:len(n_port)-1]

                    self.host = host
                    self.puerto = port
                except IndexError:
                    print('error de indice del manifest')
                    pass #si no encuentra alguna linea
        else:
            print('manifest no encontrado')
        
        if os.path.isfile(actual+ '/registry.txt'):
            with open(actual + '/registry.txt' , 'r', encoding='utf-8') as file:
                lines = file.readlines()
                try:
                    user = lines[0].split(':')
                    user = user[1]
                    user = user[:len(user)-1]

                    password = lines[1].split(':')
                    password = password[1]
                    password = password[:len(password)-1]
                    self.txt_usuario.setText(user)
                    self.txt_contra.setText(password)

                except IndexError:
                    print('error de indice del registry')
                    pass #si no encuentra alguna linea
        else:
            print('Datos de usuario no encontrados')


    def conectar(self):
        try:
            if self.host and self.puerto:
                self.conexion = rpyc.connect(self.host , self.puerto)
                self.lb_conexion.setText('CONECTADO')
            else:
                QMessageBox.about(self,'ERROR', 'Host y puerto no encontrados en el manifest' )

        except ConnectionRefusedError:
            self.lb_conexion.setText('EL SERVIDOR NO RESPONDE')
            
        except socket.error:
            self.lb_conexion.setText('SERVIDOR FUERA DE RED')

    def conectar_manual(self):
        dialog = InputDialog('HOST:','PUERTO:','CONECTAR MANUAL', self)
        dialog.resize(250,100)   
        if dialog.exec():
                hostx , puertox = dialog.getInputs()
                try:
                    if hostx != '':
                        puertox = int(puertox)
                        self.host = hostx
                        self.puerto = puertox
                        self.conexion = rpyc.connect(hostx , puertox)
                        self.lb_conexion.setText('CONECTADO')
                    else: 
                        QMessageBox.about(self,'ERROR' ,'Ingrese un host antes de continuar')
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros en el PUERTO')
                except ConnectionRefusedError:
                    self.lb_conexion.setText('EL SERVIDOR NO RESPONDE')
                    
                except socket.error:
                    self.lb_conexion.setText('SERVIDOR FUERA DE LA RED')
                    #QMessageBox.about(self,'ERROR' ,'No se puede establecer la conexion con el servidor')

    def guardar_datos(self):
        
        if self.checkBox.isChecked():

            with open(self.actual + '/registry.txt' , 'w', encoding='utf-8') as file:
                file.write('usuario:'+ self.txt_usuario.text() + '\n')
                file.write('contra:' + self.txt_contra.text()+ '\n')
                file.write('')

    def iniciar(self):
       

        if self.conexion == None: #no existe la conexion
            self.conectar()
            usuario = self.txt_usuario.text()
            contra = self.txt_contra.text()
            try:
                resultado = self.conexion.root.obtener_usuario_activo()
                encontrado = False
                for item in resultado:
                    if item[0] == usuario and item[1] == contra and item[6] == 'vendedor':
                        encontrado = True
                        self.datos_usuario = item
                        self.guardar_datos()
                        self.accept()
                if encontrado == False:
                    QMessageBox.about(self ,'ERROR', 'Usuario o contraseña invalidas')

            except EOFError:
                QMessageBox.about(self ,'Conexion', 'El servidor no responde')
            except AttributeError:
                pass

        else: #existe la conexion
            usuario = self.txt_usuario.text()
            contra = self.txt_contra.text()
            try:
                resultado = self.conexion.root.obtener_usuario_activo()
                encontrado = False
                for item in resultado:
                    if item[0] == usuario and item[1] == contra and item[6] == 'vendedor':
                        encontrado = True
                        self.datos_usuario = item
                        print('USUARIO ENCONTRADO')
                        self.guardar_datos()
                        self.accept()
                if encontrado == False:
                    QMessageBox.about(self ,'ERROR', 'Usuario o contraseña invalidas')

            except EOFError:
                QMessageBox.about(self ,'Conexion', 'El servidor no responde') 
            except AttributeError:
                pass
    
    def obt_datos(self):
        return self.datos_usuario , self.conexion ,self.host , self.puerto

    def closeEvent(self, event):
        """Generate 'question' dialog on clicking 'X' button in title bar.

        Reimplement the closeEvent() event handler to include a 'Question'
        dialog with options on how to proceed - Save, Close, Cancel buttons
        """
        reply = QMessageBox.question(
            self, "Salir",
            "¿Deseas salir de la aplicación?",
            QMessageBox.Close | QMessageBox.Cancel)

        if reply == QMessageBox.Close:
            event.accept()
        else:
            event.ignore()

class Dimensionado(QMainWindow):
    ventana_login = 0
    ventana_buscar = 0
    ventana_modificar = 0
    ventana_manual = 0
    ventana_informe = 0
    def __init__(self):
        super(Dimensionado,self).__init__()
        uic.loadUi('dimensionado1.ui',self)
        self.host = None
        self.puerto = None
        self.datos_usuario = None
        self.conexion = None
       
        self.iniciar_session()
        self.inicializar()
        self.btn_buscar.setFocus()
        self.btn_buscar.clicked.connect(self.buscar)
        self.btn_reconectar.clicked.connect(self.reconectar)
        self.btn_modificar.clicked.connect(self.modificar)
        self.btn_orden_manual.clicked.connect(self.orden_manual)
        self.btn_generar_clave.clicked.connect(self.generar_clave)
        self.btn_informe.clicked.connect(self.generar_informe)
        self.btn_atras.clicked.connect(self.atras)

    def iniciar_session(self):
        self.ventana_login = Login(self)
        self.ventana_login.show()
        salir = self.ventana_login.exec()
        if salir == 0:
            print('cerrar aplicacion')
            sys.exit(0)
        elif salir == 1:
            datos , conn , host , puerto = self.ventana_login.obt_datos()
            self.datos_usuario = datos
            self.conexion = conn
            self.host = host
            self.puerto = puerto

            print(datos)
            print('mostrar el menu')
        print('continuando..')


    def inicializar(self):
        print('inicializando...')
        self.btn_generar_clave.show()
        self.btn_orden_manual.show()
        self.btn_informe.show()

        actual = os.path.abspath(os.getcwd())
        actual = actual.replace('\\' , '/')
        ruta = actual + '/icono_imagen/madenco logo.png'
        foto = QPixmap(ruta)
        self.lb_logo.setPixmap(foto)

        self.lb_conexion.setText('CONECTADO')
        if self.datos_usuario[4] == 'NO':
            self.btn_generar_clave.hide()
            detalle = json.loads(self.datos_usuario[7])
            funciones = detalle['vendedor']
            if not 'manual' in funciones:
                self.btn_orden_manual.hide()
            if not 'informes' in funciones:
                self.btn_informe.hide()
        

        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        
        

    def buscar(self):
        self.hide()
        self.ventana_buscar = Buscar( self.conexion, self.datos_usuario ,self)
        self.ventana_buscar.show()
    
    def reconectar(self):
        try:
            self.conexion = rpyc.connect( self.host, self.puerto)
            self.lb_conexion.setText('CONECTADO')

        except ConnectionRefusedError:
            self.lb_conexion.setText('EL SERVIDOR NO RESPONDE')
            
        except socket.error:
            self.lb_conexion.setText('SERVIDOR FUERA DE RED')

    
    def modificar(self):
        self.hide()
        self.ventana_modificar = Modificar(self.conexion , self.datos_usuario, self) #v5 
        self.ventana_modificar.show()
    
    def orden_manual(self):
        if self.datos_usuario[4] == 'SI' :
            self.hide()
            self.ventana_manual = Orden_manual(self.conexion, None ,self.datos_usuario ,self)
            self.ventana_manual.show()
        else:
            dialog = InputDialog2('CLAVE:','INGRESE CLAVE',self)
            if dialog.exec():
                clave = dialog.getInputs()
                try:
                    resultado = self.conexion.root.obtener_clave()
                    end = 0
                    for item in resultado:
                        if clave == item[0]:
                            self.hide()
                            self.ventana_manual = Orden_manual(self.conexion, clave ,self.datos_usuario ,self)
                            self.ventana_manual.show()
                            end = 1
                    if end == 0:
                        QMessageBox.about(self,'ERROR' ,'CLAVE INVALIDA')

                except EOFError:
                    self.lb_conexion.setText('Se perdio la conexion con el servidor')
                    
                    #QMessageBox.about(self,'ERROR' ,'No se puede establecer la conexion con el servidor')

    def generar_clave(self):

        dialog = InputDialog2('CLAVE:', 'REGISTRAR CLAVE',self)
        if dialog.exec():
            clave = dialog.getInputs()
            if clave != '':
                try:
                    if self.conexion.root.registrar_clave(clave):
                        QMessageBox.about(self,'EXITO' ,'CLAVE: ' + clave +' REGISTRADA')
                    else:
                        QMessageBox.about(self,'ERROR' ,'ERROR AL REGISTRAR LA CLAVE')
                except EOFError:
                    QMessageBox.about(self,'ERROR' ,'Se perdio la conexion con el servidor')
            else:
                QMessageBox.about(self,'ERROR' ,'Ingrese una clave antes de continuar')

    def generar_informe(self):
        self.hide()
        self.ventana_informe = Informes(self.conexion , self)
        self.ventana_informe.show()

    def atras(self):
        self.iniciar_session()
        self.inicializar()
    
    def closeEvent(self,event):
        reply = QMessageBox.question(
            self, "Salir",
            "¿Deseas salir de la aplicación?",
            QMessageBox.Close | QMessageBox.Cancel)

        if reply == QMessageBox.Close:
            event.accept()
        else:
            event.ignore()    

class Buscar(QMainWindow):
    ventana_crear = 0

    def __init__(self, conn, usuario ,parent = None):
        super(Buscar, self).__init__(parent)
        uic.loadUi('dimensionado2.ui' , self)
        
        self.datos_usuario = usuario
        self.conexion = conn
        self.vendedores = []
        self.consultas = None
        self.aux_tabla = None
        self.tableWidget.setColumnWidth(0,80) #interno
        self.tableWidget.setColumnWidth(1,80) #documento
        self.tableWidget.setColumnWidth(2,80) #nro doc
        self.tableWidget.setColumnWidth(3,125) #fecha venta
        self.tableWidget.setColumnWidth(4,200) #vendededor
        self.tableWidget.setColumnWidth(5,100) #total

        self.radio2.setChecked(True)
        self.dateEdit.setCalendarPopup(True)
        self.dateEdit.setDate(datetime.now().date())
        self.btn_atras.clicked.connect(self.atras)
        self.btn_crear.clicked.connect(self.crear)
        self.btn_buscar.clicked.connect(self.buscar)
        self.comboBox.currentIndexChanged['QString'].connect(self.filtrar_vendedor)
        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_buscar.setIcon(QIcon('icono_imagen/buscar.ico'))
        self.btn_crear.setIcon(QIcon('icono_imagen/continuar.ico'))

    
    def buscar(self):
        largo = self.comboBox.count()
        if largo > 1:
            for i in range(largo):
                #print("borranto: " + str(largo - 1) )
                self.comboBox.removeItem(1)

        if self.conexion:

            self.tableWidget.setRowCount(0)
            if self.radio1.isChecked():  #BUSCANDO POR NUMERO INTERNO
                inter = self.txt_interno.text()
                try:
                    inter = int(inter)
                    consulta = self.conexion.root.buscar_venta_interno(inter)
                    if consulta != None :
                        print(consulta[1])
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)

                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                        if consulta[3] == 0 : #es boleta
                            self.tableWidget.setItem(fila , 1 , QTableWidgetItem( 'BOLETA' )) #TIPO DOCUMENTO
                            self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[4]) ))      #NRO DOCUMENTO
                        elif consulta[4] == 0: #es factura
                            self.tableWidget.setItem(fila , 1 , QTableWidgetItem( 'FACTURA' )) #TIPO DOCUMENTO
                            self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( consulta[3])))      #NRO DOCUMENTO

                        self.tableWidget.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                        self.tableWidget.setItem(fila , 4 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                        self.tableWidget.setItem(fila , 5 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL

                    else:
                        QMessageBox.about(self,'Busqueda' ,'Nota de venta NO encontrada')
                    
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

            elif self.radio2.isChecked(): #BUSCANDO POR FECHA ACTUAL
                lista_vendedores = []
                aux_lista = []
                date = self.dateEdit.date()
                aux = date.toPyDate()
                inicio = str(aux) + ' ' + '00:00:00'
                fin = str(aux) + ' ' + '23:59:59'
                try:
                    consultas = self.conexion.root.buscar_venta_fecha(inicio,fin)
                    if consultas != ():
                        self.consultas = consultas

                        for consulta in consultas:
                            fila = self.tableWidget.rowCount()
                            self.tableWidget.insertRow(fila)
                            self.tableWidget.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                            if consulta[3] == 0 : #es boleta
                                self.tableWidget.setItem(fila , 1 , QTableWidgetItem( 'BOLETA' )) #TIPO DOCUMENTO
                                self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[4]) ))      #NRO DOCUMENTO
                            elif consulta[4] == 0: #es factura
                                self.tableWidget.setItem(fila , 1 , QTableWidgetItem( 'FACTURA' )) #TIPO DOCUMENTO
                                self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( consulta[3])))      #NRO DOCUMENTO

                            self.tableWidget.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                            self.tableWidget.setItem(fila , 4 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                            aux = consulta[2]
                            aux = aux[0:10]
                            #print(aux)
                            
                            if aux not in aux_lista:
                                aux_lista.append(aux)
                                lista_vendedores.append( consulta[2] )
    
                            self.tableWidget.setItem(fila , 5 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL
                        
                        self.vendedores = lista_vendedores
                        for item in self.vendedores:
                            self.comboBox.addItem(item)

                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Notas de Venta')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')
            
            
            self.aux_tabla = self.tableWidget
        else:
            QMessageBox.about(self ,'Conexion', 'Se perdio la conexion')

    def rellenar_tabla(self):
        self.tableWidget.setRowCount(0)
        consultas = self.consultas
        if consultas != ():
            for consulta in consultas:
                fila = self.tableWidget.rowCount()
                self.tableWidget.insertRow(fila)
                self.tableWidget.setItem(fila , 0 , QTableWidgetItem(str(consulta[0]))) #INTERNO
                if consulta[3] == 0 : #es boleta
                    self.tableWidget.setItem(fila , 1 , QTableWidgetItem( 'BOLETA' )) #TIPO DOCUMENTO
                    self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[4]) ))      #NRO DOCUMENTO
                elif consulta[4] == 0: #es factura
                    self.tableWidget.setItem(fila , 1 , QTableWidgetItem( 'FACTURA' )) #TIPO DOCUMENTO
                    self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( consulta[3])))      #NRO DOCUMENTO

                self.tableWidget.setItem(fila , 3 , QTableWidgetItem(   str(consulta[1]  ))) #FECHA VENTA
                self.tableWidget.setItem(fila , 4 , QTableWidgetItem(    consulta[2] ))             #VENDEDOR
                self.tableWidget.setItem(fila , 5 , QTableWidgetItem(   str(consulta[5]))   )      #TOTAL
                

    def crear(self):
        
        seleccion = self.tableWidget.selectedItems()
        if seleccion != []:
            dato = seleccion[0].text()
            try:
                nro_interno = int(dato)
                self.hide()
                self.ventana_crear = Crear(self.conexion ,nro_interno, self.datos_usuario , self)
                self.ventana_crear.show()

            except ValueError:
                QMessageBox.about(self,'CONSEJO', 'Seleccione solo el nro interno')
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione un Nro interno antes de continuar')

    def filtrar_vendedor(self):
        vendedor = self.comboBox.currentText()
        
        if vendedor == 'TODOS':
            self.rellenar_tabla()
        else:
            self.rellenar_tabla()
            self.tableWidget = self.aux_tabla
            remover = []
            vendedor = vendedor[0:10]
            print('solo: ' + vendedor)
            column = 4
            # rowCount() This property holds the number of rows in the table
            for row in range(self.tableWidget.rowCount()): 
                # item(row, 0) Returns the item for the given row and column if one has been set; otherwise returns nullptr.
                _item = self.tableWidget.item(row, column) 
                if _item:            
                    item = self.tableWidget.item(row, column).text()
                    #print(f'row: {row}, column: {column}, item={item}')
                    aux_item = item[0:10]
                    if aux_item != vendedor:
                        remover.append(row)
            print(remover)

            k = 0
            for i in remover:
                self.tableWidget.removeRow(i - k)
                k += 1

    def atras(self):
        self.parent().show()
        self.hide()

class Crear(QMainWindow):
    def __init__(self, conn , interno, usuario ,parent= None):
        super(Crear, self).__init__(parent)
        uic.loadUi('dimensionado3.ui', self)

        self.datos_usuario = usuario
        self.conexion = conn
        self.tableWidget.setColumnWidth(0,80)
        self.tableWidget.setColumnWidth(1,430)
        self.tableWidget.setColumnWidth(2,85)
        self.inter = interno
        self.nro_doc= 0
        self.tipo_doc = ''
        self.fecha_venta = None
        self.vendedor = None
        self.items = None
        self.carpeta = None
        self.nro_orden = None
        self.nombre.setFocus()
        self.inicializar(interno)

        self.btn_atras.clicked.connect(self.atras)
        self.btn_agregar.clicked.connect(self.agregar)
        self.btn_eliminar.clicked.connect(self.eliminar)
        
        self.btn_registrar.clicked.connect(self.registrar)
        self.btn_rellenar.clicked.connect(self.rellenar)
        self.nombre.returnPressed.connect(self.cambiar_telefono)
        self.telefono.returnPressed.connect(self.cambiar_contacto)
        self.contacto.returnPressed.connect(self.cambiar_oce)

        #self.nombre.textChanged.connect(self.mostrar_long_nombre)
        #self.contacto.textChanged.connect(self.mostrar_long_contacto)
        #self.telefono.textChanged.connect(self.mostrar_long_telefono)

    def inicializar(self, nro_interno):
    
        actual = os.path.abspath(os.getcwd())
        self.carpeta = actual.replace('\\' , '/')
        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_agregar.setIcon(QIcon('icono_imagen/agregar.ico'))
        self.btn_eliminar.setIcon(QIcon('icono_imagen/remover.ico'))
        self.btn_registrar.setIcon(QIcon('icono_imagen/guardar.ico'))
        self.btn_rellenar.setIcon(QIcon('icono_imagen/actualizar.ico'))
        fecha = datetime.now().date()
        self.fecha.setCalendarPopup(True)
        self.fecha.setDate(fecha)
        self.lb_interno.setText(str(nro_interno))
        try:
            items = self.conexion.root.obtener_item_interno(nro_interno)
            venta = self.conexion.root.obtener_venta_interno(nro_interno) #v5 nombre obtenido

            aux = datetime.fromisoformat(str(venta[3] ) )
            self.fecha_venta = aux
            self.items = items
            self.lb_fecha.setText(str(aux.strftime("%d-%m-%Y %H:%M:%S")))
            if venta[1] == 0:
                self.tipo_doc = 'BOLETA'
                self.lb_doc.setText( self.tipo_doc )
                self.lb_documento.setText(str(venta[2]))
                self.nro_doc = venta[2]
                
            elif venta[2] == 0:
                self.tipo_doc = 'FACTURA'
                self.lb_doc.setText( self.tipo_doc )
                self.lb_documento.setText(str(venta[1]))
                self.nro_doc = venta[1]
               
            self.vendedor = venta[4]  #VENDEDOR
            if venta[5]:
                self.nombre.setText(venta[5]) #nombre del cliente factura

            self.rellenar() 

        except EOFError:
            QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')

    def rellenar(self):

        if self.items != ():
            self.tableWidget.setRowCount(0)
            for item in self.items:
                fila = self.tableWidget.rowCount()
                self.tableWidget.insertRow(fila)
                self.tableWidget.setItem(fila , 0 , QTableWidgetItem(str(item[0])) ) #CANTIDAD
                self.tableWidget.setItem(fila , 1 , QTableWidgetItem(str(item[1])) )  #DESCRIPCION
                if self.tipo_doc == 'BOLETA':
                    neto = ( item[2] / 1.19 )
                    neto = round(neto,2)  #maximo 2 decimales  continuar modificando el json de crear orden trabajo
                    self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(neto) ) )  #VALOR NETO BOLETAS VIENEN CON IVA 
                elif self.tipo_doc == 'FACTURA':
                    self.tableWidget.setItem(fila , 2 , QTableWidgetItem(str(item[2] )) )  #VALOR NETO FACTURAS VIENEN SIN IVA


    def agregar(self):
        if self.tableWidget.rowCount() <=16 :
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')

    def eliminar(self):
        fila = self.tableWidget.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tableWidget.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    
    def registrar(self):
        nombre = self.nombre.text()
        telefono = self.telefono.text()
        fecha = self.fecha.date()  #FECHA ESTIMADA
        fecha = fecha.toPyDate()
        lineas_totales = 0
        if nombre != '':
            if telefono != '':
                try:
                    telefono = int(telefono)
                    #continuar testeando con el analisis.py
                    cant = self.tableWidget.rowCount()
                    print(f'cantidad de datos: {cant}')
                    vacias = False
                    correcto = True
                    cantidades = []
                    descripciones = []
                    valores_neto = []
                    i = 0
                    while i< cant:
                        cantidad = self.tableWidget.item(i,0) #Collumna cantidades
                        descripcion = self.tableWidget.item(i,1) #Columna descripcion
                        neto = self.tableWidget.item(i,2) #Columna de valor neto
                        if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                            if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                                try: 
                                    nueva_cant = cantidad.text().replace(',','.',3)
                                    nuevo_neto = neto.text().replace(',','.',3)
                                    cantidades.append( float(nueva_cant) )
                                    descripciones.append(descripcion.text())
                                    lineas = self.separar(descripcion.text())
                                    lineas_totales = lineas_totales + len(lineas)
                                    valores_neto.append(float(nuevo_neto))

                                except ValueError:
                                    correcto = False
                            else:
                                vacias=True
                        else:
                            vacias = True
                        i+=1
                    if vacias:
                        QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
                    elif lineas_totales > 14:
                        QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
                    elif correcto == False:
                        QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
                    else:
                        formato = {
                            "cantidades" : cantidades,
                            "descripciones" : descripciones,
                            "valores_neto": valores_neto,
                            "creado_por" : self.datos_usuario[8]
                        }
                        detalle = json.dumps(formato)
                        try:
                            enchape = 'NO'
                            despacho = 'NO'

                            
                            if self.r_despacho.isChecked():
                                despacho = 'SI'
                            
                            oce = self.oce.text()
                            fecha_orden = datetime.now().date()
                            cont = self.contacto.text()

                            if self.r_dim.isChecked():

                                if self.r_enchape.isChecked():
                                    fecha = fecha + timedelta(days=2)
                                    enchape = 'SI'

                                self.conexion.root.registrar_orden_dimensionado( self.inter , str(self.fecha_venta), nombre , telefono, str(fecha) , detalle, self.tipo_doc, self.nro_doc,enchape,despacho,str(fecha_orden),cont,oce,self.vendedor )
                                resultado = self.conexion.root.buscar_orden_dim_interno(self.inter)
                                self.nro_orden = self.buscar_nro_orden(resultado)
                                datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, enchape, cont,oce,self.vendedor)
                                self.crear_pdf(datos,'dimensionado')
                                boton = QMessageBox.question(self, 'Orden de dimensionado registrada correctamente', 'Desea abrir la Orden?')
                                if boton == QMessageBox.Yes:
                                    self.ver_pdf('dimensionado')
                            elif self.r_elab.isChecked():
                                self.conexion.root.registrar_orden_elaboracion( nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)
                                resultado = self.conexion.root.buscar_orden_elab_interno(self.inter)
                                self.nro_orden = self.buscar_nro_orden(resultado)

                                datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, 'NO', cont, oce, self.vendedor)
                                self.crear_pdf(datos , 'elaboracion')
                                boton = QMessageBox.question(self, 'Orden de elaboracion registrada correctamente', 'Desea abrir la Orden?')
                                if boton == QMessageBox.Yes:
                                    self.ver_pdf('elaboracion')
                            elif self.r_carp.isChecked():
                                self.conexion.root.registrar_orden_carpinteria( nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)
                                resultado = self.conexion.root.buscar_orden_carp_interno(self.inter)
                                self.nro_orden = self.buscar_nro_orden(resultado)

                                datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, 'NO', cont, oce, self.vendedor)
                                self.crear_pdf(datos , 'carpinteria')
                                boton = QMessageBox.question(self, 'Orden de elaboracion registrada correctamente', 'Desea abrir la Orden?')
                                if boton == QMessageBox.Yes:
                                    self.ver_pdf('carpinteria')
                            elif self.r_pall.isChecked():
                                self.conexion.root.registrar_orden_pallets( nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, self.inter ,detalle, str(self.fecha_venta), self.vendedor)
                                resultado = self.conexion.root.buscar_orden_pall_interno(self.inter)
                                self.nro_orden = self.buscar_nro_orden(resultado)

                                datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, 'NO', cont, oce, self.vendedor)
                                self.crear_pdf( datos , 'pallets')
                                boton = QMessageBox.question(self, 'Orden de elaboracion registrada correctamente', 'Desea abrir la Orden?')
                                if boton == QMessageBox.Yes:
                                    self.ver_pdf('pallets')
                            else:
                                QMessageBox.about(self, 'ERROR', 'Seleccione un tipo de orden a generar, antes de proceder a registrar')    

                        except EOFError:
                            print('Se perdio la conexion con el servidor')
                        except AttributeError:
                            QMessageBox.about(self,'IMPORTANTE', 'Este mensaje se debe a que hubo un error al ingresar los datos a la base de datos. Contacte con el soporte')


                except ValueError:
                    QMessageBox.about(self, 'ERROR', 'Solo ingrese numeros en el campo "Telefono" ')          
            else:
                QMessageBox.about(self, 'Sugerencia', 'Ingrese un telefono antes de continuar')           
        else:
            QMessageBox.about(self, 'Sugerencia', 'Ingrese un nombre antes de continuar')
  
    def crear_pdf(self, lista, tipo):
        ruta = ( self.carpeta +'/ordenes/' + tipo +'_' + lista[0] + '.pdf' )  #NRO DE ORDEN  
        formato = self.carpeta +"/formatos/" + tipo +".jpg"
        agua = self.carpeta + "/formatos/despacho.png"
        hojas = 2
        if tipo == 'carpinteria':
            hojas = 1
        try:
            documento = canvas.Canvas(ruta)

            for pagina in range(hojas):
                documento.setPageSize(( 216 * mm , 279 * mm))
                documento.drawImage( formato, 0* mm , 2 * mm , 216 *mm ,279 *mm )
                if self.r_despacho.isChecked():
                    documento.setFillAlpha(0.6)
                    documento.drawImage( agua , 83* mm , 30* mm , 100*mm ,100*mm , mask= 'auto')
                    documento.drawImage( agua , 83* mm , (30+136)* mm , 100*mm ,100*mm , mask= 'auto')

                documento.setFillAlpha(1)
                documento.drawString( 0 * mm, 139.5 * mm ,'------------------------------------------------------------------------------------------')
                documento.drawString( 105 * mm, 139.5 * mm ,'----------------------------------------------------------------------------------------------')

                documento.rotate(90)

                documento.setFont('Helvetica',10)

                k = 2.5 #constante
                salto = 0
                for i in range(2):
                    documento.setFont('Helvetica',9)
                    documento.drawString( (28 + k + salto) *mm , -59.5 * mm , lista[2] )  #NOMBRE

                    documento.drawString( (100 + k + salto) *mm , -66 * mm , str(lista[3]) )   #TELEFONO
                    documento.drawString( (106 + k +salto) *mm , -59.5 * mm , lista[1]  )    #FECHA DE ORDEN

                    
                    
                    if self.tipo_doc == 'FACTURA':
                        documento.drawString( (45 + k + salto) *mm , -85 * mm , str(self.nro_doc) )     #NRO FACTURA
                    elif self.tipo_doc == 'BOLETA':
                        documento.drawString( (15 + k + salto) *mm , -85 * mm ,  str(self.nro_doc) )      #NRO BOLETA

                    if tipo == 'dimensionado':
                        documento.drawString( (88 + k + salto) *mm , -94 * mm ,  lista[7] )   #ENCHAPE

                    documento.drawString( (110 + k + salto) *mm , -85 * mm ,   lista[9]  )       #ORDEN DE COMPRA

                    documento.drawString( (32 + k + salto) *mm , -66 * mm ,  lista[8] )   #CONTACTO

                    documento.drawString( (33+ salto) *mm , -205 * mm , lista[10] ) #NOMBRE VENDEDOR
                    
                    documento.setFont('Helvetica-Bold',12)
                    documento.drawString( (106 + k + salto) *mm , -44.5 * mm , lista[4]  ) #FECHA ESTIMADA
                    documento.drawString( (106 + k + salto) *mm , -20.5 * mm , lista[0]  ) #NRO DE ORDEN

                    documento.setFillColor(white)
                    documento.setStrokeColor(white)
                    documento.rect((5+ k + salto )* mm, -200 * mm ,80*mm,3*mm, fill=1 )
                    documento.setFillColor(black)
                    documento.setStrokeColor(black)
                    documento.setFont('Helvetica-Bold',10)

                    if pagina == 0:
                        if k == 2.5:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'ORIGINAL' )
                        else:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA ' + tipo.upper() )
                    else:
                        if k == 2.5:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA CLIENTE' )
                        else:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA BODEGA' )
                    documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )        
                    salto += 139.5
                    k = 0 

                #items
                cons = 108
                documento.setFont('Helvetica',9)
                align_cant_1 = 12
                align_descr_1 = 22
                align_cant_2 = 151
                align_descr_2 = 161

                cantidades = lista[5]
                descripciones = lista[6]
                i = 0
                while i < len(cantidades):
                    documento.drawString(align_cant_1 *mm , -cons* mm , str(cantidades[i]) )  #cantidad
                    
                    documento.drawString(align_cant_2 *mm , -cons * mm , str(cantidades[i]) )

                    cadenas = self.separar(descripciones[i])

                    for cadena in cadenas:
                        documento.drawString(align_descr_1 *mm , -cons* mm , cadena )  #descripcion

                        documento.drawString(align_descr_2 *mm , -cons* mm , cadena ) #54 LONG MAXIMA
                        cons += 5
                    i +=1

                documento.showPage()
            #documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )
            documento.save()
            sleep(1)
        except PermissionError:
            QMessageBox.about(self,'ERROR', 'Otro programa esta modificando este archivo por lo cual no se puede modificar actualmente.')

        
    def ver_pdf(self, tipo):

        abrir = self.carpeta+ '/ordenes/' + tipo +'_' +str(self.nro_orden) + '.pdf'
        subprocess.Popen([abrir], shell=True)

    def buscar_nro_orden(self,tupla):
        mayor = 0
        for item in tupla:
            if item[0] > mayor :
                mayor = item[0]
        return mayor
    
    def separar(self,cadena):
        lista = []
        iter = len(cadena)/54
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
       # print(cadena)
        #print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> 54:
        
            #print('long > 54:')
            aux = cadena[0:54]
            index = aux[::-1].find(' ')
        
            aux = aux[:(54-index)]
           # print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[54 - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                print('item O CADENA vacio')
                #print('----------------------------------------------------')
            else:
               # print('LONG < 54: ' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista

    def cambiar_telefono(self):
        self.telefono.setFocus()
    def cambiar_contacto(self):
        self.contacto.setFocus()
    def cambiar_oce(self):
        self.oce.setFocus()
    
    '''def mostrar_long_nombre(self):
        nombre = self.nombre.text()  #31 MAX SOLO CARACTERES
        vacios = nombre.count(' ')  #3 espacios = 1 caracter
        razon = int(vacios/2) 

        #print('espacios: ' + str( vacios ) )

        if ((len(nombre) - vacios) + razon) <= 31: 
            print('TAMAÑO NOMBRE VALIDO')
        else:
            print('tamaño nombre superado')
    
        self.max_nombre.setText( str(len( self.nombre.text() )))

    def mostrar_long_contacto(self):
        self.max_contacto.setText( str(len( self.contacto.text() )) )
    def mostrar_long_telefono(self):
        self.max_telefono.setText( str(len( self.telefono.text() )) ) '''
    
    def atras(self):
        self.parent().show()
        self.hide()

class Modificar(QMainWindow):
    ventana_datos = 0
    ventana_reingreso = 0
    def __init__(self, conn, dato_usuario ,parent = None):
        super(Modificar , self ).__init__(parent)
        uic.loadUi('dimensionado4.ui', self)
        self.datos_usuario = dato_usuario
        self.conexion = conn
        self.r_fecha.setChecked(True)
        self.tableWidget.setColumnWidth(0,70)
        self.tableWidget.setColumnWidth(1,70)
        self.tableWidget.setColumnWidth(2,100)
        self.tableWidget.setColumnWidth(3,170)
        self.tableWidget.setColumnWidth(4,120)
        self.tableWidget.setColumnWidth(5,100)
        self.tableWidget.setColumnWidth(6,100)
        self.dateEdit.setCalendarPopup(True)
        self.dateEdit.setDate(datetime.now().date())
        self.seleccion = None

        self.btn_modificar.clicked.connect(self.modificar)
        self.btn_atras.clicked.connect(self.atras)
        self.btn_dimensionado.clicked.connect(self.buscar_dimensionado)
        self.btn_elaboracion.clicked.connect(self.buscar_elaboracion)
        self.btn_carpinteria.clicked.connect(self.buscar_carpinteria)
        self.btn_pallets.clicked.connect(self.buscar_pallets)
        self.btn_reingreso.clicked.connect(self.reingreso)
        self.btn_ver.clicked.connect(self.ver_pdf)

        self.ch_nulas.stateChanged.connect(self.solo_nulas)

        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_dimensionado.setIcon(QIcon('icono_imagen/buscar.ico'))
        self.btn_elaboracion.setIcon(QIcon('icono_imagen/buscar.ico'))
        self.btn_carpinteria.setIcon(QIcon('icono_imagen/buscar.ico'))
        self.btn_pallets.setIcon(QIcon('icono_imagen/buscar.ico'))
        self.btn_modificar.setIcon(QIcon('icono_imagen/editar.ico'))
        self.btn_reingreso.setIcon(QIcon('icono_imagen/continuar.ico'))

    def buscar_dimensionado(self):
        self.ch_nulas.setChecked(False)
        self.seleccion = 'DIMENSIONADO'
        self.lb_tipo_orden.setText(self.seleccion)
        if self.conexion:

            self.tableWidget.setRowCount(0)

            if self.r_orden.isChecked(): #Busqueda por numero interno

                orden = self.txt_orden.text()
                try:
                    orden = int(orden)
                    consulta = self.conexion.root.buscar_orden_dim_numero(orden)
                    if consulta != None :
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(consulta[0])) )  #nro orden
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(consulta[1]) + ' "'  ))  #nro interno
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[11]) ) )       #fecha creacion
                        self.tableWidget.setItem(fila , 3 , QTableWidgetItem(consulta[3]))       #nombre
                        self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(consulta[2]) ) )       #fecha venta
                        self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(consulta[5]) ) )       #fecha estimada
                        if consulta[19]:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'ANULADA' ) )       #estado
                        else:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #estado
                    else:
                        QMessageBox.about(self,'Busqueda' ,'Orden de dimensionado NO encontrada')
                    
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

            elif self.r_fecha.isChecked():
                date = self.dateEdit.date()
                aux = date.toPyDate()
                #inicio = str(aux) + ' ' + '00:00:00'
                #fin = str(aux) + ' ' + '23:59:59'    
                try:
                    datos = self.conexion.root.buscar_orden_dim_fecha( str(aux) )
                    if datos != ():
                        for dato in datos:
                            fila = self.tableWidget.rowCount()
                            self.tableWidget.insertRow(fila)
                            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(dato[0])) )  #nro orden
                            self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(dato[1]) + ' "'  ))  #nro interno
                            self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(dato[4]) ) ) #fecha creacion
                            self.tableWidget.setItem(fila , 3 , QTableWidgetItem(dato[3]))       #nombre
                            self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(dato[2]) ) )       #fecha venta
                            self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(dato[5]) ) )       #fecha estimada
                            if dato[6]:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem('ANULADA') )       #estado
                            else:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #estado
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de dimensionado')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

        else:
            QMessageBox.about(self ,'Conexion', 'Se perdio la conexion')
        

    def buscar_elaboracion(self):
        self.ch_nulas.setChecked(False)
        self.seleccion = 'ELABORACION'
        self.lb_tipo_orden.setText(self.seleccion)
        if self.conexion:
    
            self.tableWidget.setRowCount(0)

            if self.r_orden.isChecked(): #Busqueda por numero de orden

                orden = self.txt_orden.text()
                try:
                    orden = int(orden)
                    consulta = self.conexion.root.buscar_orden_elab_numero(orden)
                    if consulta != None :
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(consulta[0])) )  #nro orden
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(consulta[10]) + ' "'  ))  #nro interno
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[3]) ) )       #fecha creacion
                        self.tableWidget.setItem(fila , 3 , QTableWidgetItem(consulta[1]))       #nombre
                        self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(consulta[12]) ) )       #fecha venta
                        self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(consulta[4]) ) )       #fecha estimada
                        if consulta[16]:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem('ANULADA' ) )       #fecha real
                        else:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #fecha real
                    else:
                        QMessageBox.about(self,'Busqueda' ,'Orden de elaboracion NO encontrada')
                    
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

            elif self.r_fecha.isChecked():
                date = self.dateEdit.date()
                aux = date.toPyDate()
                #inicio = str(aux) + ' ' + '00:00:00'
                #fin = str(aux) + ' ' + '23:59:59'    
                try:
                    datos = self.conexion.root.buscar_orden_elab_fecha( str(aux) )
                    if datos != ():
                        for dato in datos:
                            fila = self.tableWidget.rowCount()
                            self.tableWidget.insertRow(fila)
                            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(dato[0])) )  #nro orden
                            self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(dato[1]) + ' "'  ))  #nro interno
                            self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(dato[2]) ) ) #fecha creacion
                            self.tableWidget.setItem(fila , 3 , QTableWidgetItem(dato[3]))       #nombre
                            self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(dato[4]) ) )       #fecha venta
                            self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(dato[5]) ) )       #fecha estimada
                            if dato[6]:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'ANULADA' ) )       #estado
                            else:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #estado
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de elaboracion')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

        else:
            QMessageBox.about(self ,'Conexion', 'Se perdio la conexion')
    
    def buscar_carpinteria(self):
        self.ch_nulas.setChecked(False)
        self.seleccion = 'CARPINTERIA' 
        self.lb_tipo_orden.setText(self.seleccion)
        if self.conexion:
        
            self.tableWidget.setRowCount(0)

            if self.r_orden.isChecked(): #Busqueda por numero de orden

                orden = self.txt_orden.text()
                try:
                    orden = int(orden)
                    consulta = self.conexion.root.buscar_orden_carp_numero(orden)
                    if consulta != None :
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(consulta[0])) )  #nro orden
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(consulta[10]) + ' "'  ))  #nro interno
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[3]) ) )       #fecha creacion
                        self.tableWidget.setItem(fila , 3 , QTableWidgetItem(consulta[1]))       #nombre
                        self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(consulta[12]) ) )       #fecha venta
                        self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(consulta[4]) ) )       #fecha estimada
                        if consulta[16]:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem('ANULADA' ) )       #fecha real
                        else:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #fecha real
                    else:
                        QMessageBox.about(self,'Busqueda' ,'Orden de carpinteria NO encontrada')
                    
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

            elif self.r_fecha.isChecked():
                date = self.dateEdit.date()
                aux = date.toPyDate()
                #inicio = str(aux) + ' ' + '00:00:00'
                #fin = str(aux) + ' ' + '23:59:59'    
                try:
                    datos = self.conexion.root.buscar_orden_carp_fecha( str(aux) )
                    if datos != ():
                        for dato in datos:
                            fila = self.tableWidget.rowCount()
                            self.tableWidget.insertRow(fila)
                            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(dato[0])) )  #nro orden
                            self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(dato[1]) + ' "'  ))  #nro interno
                            self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(dato[2]) ) ) #fecha creacion
                            self.tableWidget.setItem(fila , 3 , QTableWidgetItem(dato[3]))       #nombre
                            self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(dato[4]) ) )       #fecha venta
                            self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(dato[5]) ) )       #fecha estimada
                            if dato[6]:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'ANULADA' ) )       #fecha real
                            else:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #fecha real
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de carpinteria')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

        else:
            QMessageBox.about(self ,'Conexion', 'Se perdio la conexion')
    
    def buscar_pallets(self):
        self.ch_nulas.setChecked(False)
        self.seleccion = 'PALLETS'
        self.lb_tipo_orden.setText(self.seleccion)
        if self.conexion:
        
            self.tableWidget.setRowCount(0)

            if self.r_orden.isChecked(): #Busqueda por numero de orden

                orden = self.txt_orden.text()
                try:
                    orden = int(orden)
                    consulta = self.conexion.root.buscar_orden_pall_numero(orden)
                    if consulta != None :
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(consulta[0])) )  #nro orden
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(consulta[10]) + ' "'  ))  #nro interno
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(consulta[3]) ) )       #fecha creacion
                        self.tableWidget.setItem(fila , 3 , QTableWidgetItem(consulta[1]))       #nombre
                        self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(consulta[12]) ) )       #fecha venta
                        self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(consulta[4]) ) )       #fecha estimada
                        if consulta[16]:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem('ANULADA' ) )       #fecha real
                        else:
                            self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #fecha real
                    else:
                        QMessageBox.about(self,'Busqueda' ,'Orden de pallets NO encontrada')
                    
                except ValueError:
                    QMessageBox.about(self,'ERROR' ,'Ingrese solo numeros')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

            elif self.r_fecha.isChecked():
                date = self.dateEdit.date()
                aux = date.toPyDate()
                #inicio = str(aux) + ' ' + '00:00:00'
                #fin = str(aux) + ' ' + '23:59:59'    
                try:
                    datos = self.conexion.root.buscar_orden_pall_fecha( str(aux) )
                    if datos != ():
                        for dato in datos:
                            fila = self.tableWidget.rowCount()
                            self.tableWidget.insertRow(fila)
                            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str(dato[0])) )  #nro orden
                            self.tableWidget.setItem(fila , 1 , QTableWidgetItem('" ' + str(dato[1]) + ' "'  ))  #nro interno
                            self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str(dato[2]) ) ) #fecha creacion
                            self.tableWidget.setItem(fila , 3 , QTableWidgetItem(dato[3]))       #nombre
                            self.tableWidget.setItem(fila , 4 , QTableWidgetItem( str(dato[4]) ) )       #fecha venta
                            self.tableWidget.setItem(fila , 5 , QTableWidgetItem( str(dato[5]) ) )       #fecha estimada
                            if dato[6]:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'ANULADA' ) )       #fecha real
                            else:
                                self.tableWidget.setItem(fila , 6 , QTableWidgetItem( 'VALIDA' ) )       #fecha real
                    else:
                        QMessageBox.about(self ,'Resultado', 'No se encontraron Ordenes de pallets')
                except EOFError:
                    QMessageBox.about(self,'ERROR','Se perdio la conexion con el servidor')

        else:
            QMessageBox.about(self ,'Conexion', 'Se perdio la conexion')

    def modificar(self):
        seleccion = self.tableWidget.selectedItems()
        if seleccion != []:
            dato = seleccion[0].text()
            try:
                nro_orden = int(dato)
                self.hide()
                self.ventana_datos = Guardar_cambios(self.conexion , nro_orden, self.seleccion, self.datos_usuario , self)
                self.ventana_datos.show()

            except ValueError:
                QMessageBox.about(self,'CONSEJO', 'Seleccione solo el Numero de orden')
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione un Numero de orden antes de continuar')

    def reingreso(self):
        seleccion = self.tableWidget.selectedItems()
        if seleccion != []:
            dato = seleccion[0].text()
            try:
                nro_orden = int(dato)
                self.hide()
                self.ventana_reingreso = Reingreso(self.conexion , nro_orden , self.seleccion, self)
                self.ventana_reingreso.show()

            except ValueError:
                QMessageBox.about(self,'CONSEJO', 'Seleccione solo el Numero de orden')
        else:
            QMessageBox.about(self,'ERROR', 'Seleccione un Numero de orden antes de continuar')

    def ver_pdf(self):
        seleccion = self.tableWidget.selectedItems()
        actual = os.path.abspath(os.getcwd())
        actual = actual.replace('\\' , '/')
        actual = actual + '/ordenes/'
        if seleccion != []:
            dato = seleccion[0].text()
            try:
                nro_orden = int(dato)

                if self.seleccion == 'DIMENSIONADO' :
                    abrir = actual + 'dimensionado_' +str(nro_orden) + '.pdf'
                    
                elif self.seleccion == 'ELABORACION' :
                    abrir = actual + 'elaboracion_' + str(nro_orden)+ '.pdf'

                elif self.seleccion == 'CARPINTERIA' :
                    abrir = actual + 'carpinteria_' + str(nro_orden)+ '.pdf'
                elif self.seleccion == 'PALLETS' :
                    abrir = actual + 'pallets_' + str(nro_orden)+ '.pdf'

                if os.path.isfile(abrir):
                    subprocess.Popen([abrir], shell=True)
                else:

                    QMessageBox.about(self, 'Error', 'No se encontro la orden de dimensionado en la carpeta "ordenes". Para ver el PDF seleccione el boton "Modificar orden" y luego en la ventana seleccione el boton "ver PDF"')
                
            except ValueError:
                    QMessageBox.about(self,'CONSEJO', 'Seleccione solo el Numero de orden')
            
        else:
            QMessageBox.about(self,'CONSEJO', 'Seleccione un Numero de orden para poder visualizar el PDF.')

    def vaciar_tabla(self):
        self.tableWidget.setRowCount(0)
    def solo_nulas(self):
    
        if self.ch_nulas.isChecked():
            remover = []
            print('solo nulas')
            column = 6
            # rowCount() This property holds the number of rows in the table
            for row in range(self.tableWidget.rowCount()): 
                # item(row, 0) Returns the item for the given row and column if one has been set; otherwise returns nullptr.
                _item = self.tableWidget.item(row, column) 
                if _item:            
                    item = self.tableWidget.item(row, column).text()
                    print(f'row: {row}, column: {column}, item={item}')
                    if item == 'VALIDA':
                        remover.append(row)
            print(remover)

            k = 0
            for i in remover:
                self.tableWidget.removeRow(i - k)
                k += 1
            
        else:
            if self.seleccion == 'DIMENSIONADO':
                self.buscar_dimensionado()
            elif self.seleccion == 'ELABORACION':
                self.buscar_elaboracion()
            elif self.seleccion == 'CARPINTERIA':
                self.buscar_carpinteria()
            elif self.seleccion == 'PALLETS':
                self.buscar_pallets()

    def atras(self):
        self.hide()
        self.parent().show()

class Guardar_cambios(QMainWindow):
    def __init__(self, conn , orden, selecc, dato_user ,parent= None):
        super(Guardar_cambios, self).__init__(parent)
        uic.loadUi('dimensionado5.ui', self)
        self.datos_usuario = dato_user
        self.conexion = conn
        self.seleccion = selecc
        self.tableWidget.setColumnWidth(0,80)
        self.tableWidget.setColumnWidth(1,430)
        self.tableWidget.setColumnWidth(2,85)
        self.nro_orden = orden
        self.fecha_venta = None
        self.fecha_orden = None
        self.interno = None
        self.carpeta = None
        self.vendedor = None
        self.manual = False
        self.nro_doc = None
        self.tipo_doc = None
        self.inicializar()
        
        self.btn_atras.clicked.connect(self.atras)
        self.btn_agregar.clicked.connect(self.agregar)
        self.btn_eliminar.clicked.connect(self.eliminar)
        self.btn_guardar.clicked.connect(self.guardar)
        self.btn_ver.clicked.connect(self.ver_pdf)
        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_agregar.setIcon(QIcon('icono_imagen/agregar.ico'))
        self.btn_eliminar.setIcon(QIcon('icono_imagen/remover.ico'))
        self.btn_guardar.setIcon(QIcon('icono_imagen/guardar.ico'))
        self.btn_anular.clicked.connect(self.anular)
        

    def inicializar(self):
        self.date_venta.setCalendarPopup(True)
        self.fecha.setCalendarPopup(True)
        actual = os.path.abspath(os.getcwd())
        self.carpeta = actual.replace('\\' , '/')
        cantidades = None
        descripciones = None
        enchapado = 'NO'
        self.date_venta.setDate( datetime.now())
        print('llego')

        if self.seleccion == 'DIMENSIONADO':
            try:
                resultado = self.conexion.root.buscar_orden_dim_numero(self.nro_orden)
                if resultado != None :

                    if resultado[18]: #si es manual
                        self.manual = True
                        if resultado[8]:
                            self.txt_nro_doc.setText( str(resultado[8]) )       #NUMERO DOCUMENTO
                            self.nro_doc = int( resultado[8] )

                        else:
                            self.txt_nro_doc.setText( '0' )       #NUMERO DOCUMENTO
                        if resultado[7]:
                            if resultado[7] == 'BOLETA':
                                self.comboBox.addItem( resultado[7] )                 #TIPO DOCUMENTO
                                self.comboBox.addItem('FACTURA')
                                self.tipo_doc = resultado[7]

                            elif resultado[6] == 'FACTURA':
                                self.comboBox.addItem( resultado[7] )                 #TIPO DOCUMENTO
                                self.comboBox.addItem('BOLETA')
                                self.tipo_doc = resultado[7]
                        else:
                            self.comboBox.addItem('NO ASIGNADO') 
                            self.comboBox.addItem('FACTURA')  
                            self.comboBox.addItem('BOLETA') 
                        if resultado[2]:
                            aux3 = datetime.fromisoformat(str(resultado[2]) )
                            self.date_venta.setDate( aux3 )   #FECHA DE VENTA
                        if resultado[15]:
                            self.vendedor = resultado[14]   #VENDEDOR
                            self.txt_vendedor.setText( resultado[15] )             
                    else: #SI NO ES MANUAL
                        self.manual = False
                        self.txt_interno.setEnabled(False) #solo lectura
                        self.txt_vendedor.setEnabled(False)
                        self.txt_nro_doc.setEnabled(False)
                        self.date_venta.setEnabled(False)
                        self.comboBox.setEnabled(False)

                        self.comboBox.addItem( resultado[7] )    #tipo documento
                        aux1 = datetime.fromisoformat(str(resultado[2]) )
                        self.date_venta.setDate( aux1 )     #fecha de venta
                        self.txt_nro_doc.setText( str(resultado[8]) ) #numero de documento
                        self.vendedor = resultado[15]    #vendedor
                        self.txt_vendedor.setText( resultado[15] ) #vendedor

                    self.txt_interno.setText(str( resultado[1] )) #INTERNO
                    self.interno = str( resultado[1] )
                    
                    self.nombre.setText( resultado[3] )   #nombre
                    self.telefono.setText( str(resultado[4]) ) #telefono
                    aux = datetime.fromisoformat(str(resultado[5]) )
                    self.fecha.setDate( aux )  #FECHA ESTIMADA           
                    detalle = json.loads(resultado[6])
                    
                    #self.lb_planchas.setText( str(detalle["total_planchas"]) )
                    cantidades = detalle["cantidades"]
                    descripciones = detalle["descripciones"]
                    valores_neto = detalle["valores_neto"]
                    j = 0
                    while j < len( cantidades ):
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str( cantidades[j] )) ) 
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem( descripciones[j] ) )
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( valores_neto[j] )) ) 
                        j+=1      
                    if resultado[9] == 'SI' :
                        self.r_enchape.setChecked(True)      #enchape
                        enchapado = 'SI'
                    if resultado[10] == 'SI' :
                        self.r_despacho.setChecked(True)  #despacho

                    self.lb_fecha_orden.setText( str(resultado[11]))
                    self.fecha_orden = datetime.fromisoformat( str(resultado[11]) )  #FEcha orden
                    self.contacto.setText( resultado[12] ) #contacto
                    self.oce.setText( resultado[13] )      #orden comprar
                    

            except EOFError:
                QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')

        else: 
            try:
                if self.seleccion == 'ELABORACION':
                    print('buscando elaboracion datos...')
                    resultado = self.conexion.root.buscar_orden_elab_numero(self.nro_orden)
                elif self.seleccion == 'CARPINTERIA':
                    print('buscando carpinteria datos...')
                    resultado = self.conexion.root.buscar_orden_carp_numero(self.nro_orden)
                elif self.seleccion == 'PALLETS':
                    print('buscando pallets datos...')
                    resultado = self.conexion.root.buscar_orden_pall_numero(self.nro_orden)
                
                
                if resultado != None :
                    if resultado[ 15 ]: #si es manual
                        self.manual = True
                        if resultado[5]:
                            self.txt_nro_doc.setText( str(resultado[5]) )       #NUMERO DOCUMENTO
                            self.nro_doc = int( resultado[5] )
                        else:
                            self.txt_nro_doc.setText( '0' )       #NUMERO DOCUMENTO

                        if resultado[6]:
                            if resultado[6] == 'BOLETA':
                                self.comboBox.addItem( resultado[6] )                 #TIPO DOCUMENTO
                                self.comboBox.addItem('FACTURA')
                                self.tipo_doc = resultado[6]
                            elif resultado[6] == 'FACTURA':
                                self.comboBox.addItem( resultado[6] )                 #TIPO DOCUMENTO
                                self.comboBox.addItem('BOLETA')
                                self.tipo_doc = resultado[6]
                        else:
                            self.comboBox.addItem('NO ASIGNADO')  
                            self.comboBox.addItem('FACTURA')  
                            self.comboBox.addItem('BOLETA') 
                        if resultado[12]:
                            aux3 = datetime.fromisoformat(str(resultado[12]) )
                            self.date_venta.setDate( aux3 )   #FECHA DE VENTA
                        if resultado[14]:
                            self.vendedor = resultado[14]   #VENDEDOR
                            self.txt_vendedor.setText( resultado[14] )

                    else: #si no es manual
                        self.manual = False
                        self.txt_interno.setEnabled(False) #solo lectura
                        self.txt_vendedor.setEnabled(False)
                        self.txt_nro_doc.setEnabled(False)
                        self.date_venta.setEnabled(False)
                        self.comboBox.setEnabled(False)

                        self.txt_nro_doc.setText( str(resultado[5]) )       #NUMERO DOCUMENTO
                        self.comboBox.addItem( resultado[6] )                 #TIPO DOCUMENTO
                        aux3 = datetime.fromisoformat(str(resultado[12]) )
                        self.date_venta.setDate( aux3 )   #FECHA DE VENTA
                        self.vendedor = resultado[14]   #VENDEDOR
                        self.txt_vendedor.setText( resultado[14] )

                    self.txt_interno.setText(str( resultado[10] ))        #NRO INTERNO
                    self.interno = str( resultado[10] )
                    self.r_enchape.setVisible(False)
                    self.nombre.setText( resultado[1] )         #NOMBRE CLIENTE
                    self.telefono.setText( str(resultado[2]) )  #TELEFONO

                    self.fecha_orden = datetime.fromisoformat( str(resultado[3]) )  
                    self.lb_fecha_orden.setText( str(resultado[3]))   #FECHA DE ORDEN
                    
                    f_estimada = datetime.fromisoformat(str(resultado[4]) )
                    self.fecha.setDate( f_estimada )                   #FECHA ESTIMADA DE ENTREGA

                    
                    self.contacto.setText( resultado[7] )             #CONTACTO
                    self.oce.setText( resultado[8] )                   #ORDEN DE COMPRA
                              
                    if resultado[9] == 'SI' :
                        self.r_despacho.setChecked(True)                #DESPACHO


                    detalle = json.loads(resultado[11])                  #DETALLE
                    cantidades = detalle["cantidades"]
                    descripciones = detalle["descripciones"]
                    valores_neto = detalle["valores_neto"]
                    j = 0
                    while j < len( cantidades ):
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( str( cantidades[j] )) ) 
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem( descripciones[j] ) )
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( valores_neto[j] )) ) 
                        j+=1    
            except EOFError:
                QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')

        tipo = None
        if self.seleccion == "DIMENSIONADO":
            abrir = self.carpeta + '/ordenes/' + 'dimensionado_' +str(self.nro_orden) + '.pdf'
            tipo = 'dimensionado'
        elif self.seleccion == "ELABORACION":
            abrir = self.carpeta + '/ordenes/'  + 'elaboracion_' +str(self.nro_orden) + '.pdf'
            tipo = 'elaboracion'
        elif self.seleccion == "CARPINTERIA":
            abrir = self.carpeta + '/ordenes/'  + 'carpinteria_' +str(self.nro_orden) + '.pdf'
            tipo = 'carpinteria'
        elif self.seleccion == "PALLETS":
            abrir = self.carpeta + '/ordenes/'  + 'pallets_' +str(self.nro_orden) + '.pdf'
            tipo = 'pallets'
        print('verifica el pdf')
        if not os.path.isfile(abrir):
            datos = ( str(self.nro_orden) ,str(self.fecha_orden.strftime("%d-%m-%Y")),self.nombre.text(),self.telefono.text(), str((self.fecha.date().toPyDate()).strftime("%d-%m-%Y")),cantidades,descripciones,enchapado ,self.contacto.text(),self.oce.text() )
            self.actualizar_pdf(datos, tipo)
            print('pdf no encontrado, pero se acaba de crear')
        else:
            print('El pdf si existe localmente')
        
            #( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado,contacto,oce)
            
    def agregar(self):
        fila = self.tableWidget.rowCount()
        self.tableWidget.insertRow(fila)

    def eliminar(self):
        fila = self.tableWidget.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            print('Eliminando la fila ' + str(fila))
            self.tableWidget.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    
    def guardar(self):
        nombre = self.nombre.text()
        telefono = self.telefono.text()
        contacto = self.contacto.text()
        oce = self.oce.text()
        fecha = self.fecha.date()  #fecha estimada
        fecha = fecha.toPyDate()
        enchapado = 'NO'
        despacho = 'NO'

        self.tipo_doc = self.comboBox.currentText()
        self.vendedor = self.txt_vendedor.text()
        fecha_venta = str( self.date_venta.date().toPyDate() )
        lineas_totales = 0 
        if nombre != '':
            if telefono != '':
                try:
                    telefono = int(telefono)
                    self.interno = int( self.txt_interno.text() )
                    self.nro_doc = int( self.txt_nro_doc.text() )
                    
                    cant = self.tableWidget.rowCount() #cantidad de items de la tabla
                    vacias = False
                    correcto = True
                    cantidades = []
                    descripciones = []
                    valores_neto = []
                    i = 0
                    while i< cant:
                        cantidad = self.tableWidget.item(i,0) #Collumna cantidades
                        descripcion = self.tableWidget.item(i,1) #Columna descripcion
                        neto = self.tableWidget.item(i,2) #Columna de valor neto
                        if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                            if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                                try: 
                                    nueva_cant = cantidad.text().replace(',','.',3)
                                    nuevo_neto = neto.text().replace(',','.',3)
                                    cantidades.append( float(nueva_cant) )
                                    descripciones.append(descripcion.text())
                                    lineas = self.separar(descripcion.text())
                                    lineas_totales = lineas_totales + len(lineas)
                                    valores_neto.append(float(nuevo_neto))

                                except ValueError:
                                    correcto = False
                            else:
                                vacias=True
                        else:
                            vacias = True
                        i+=1
                    if vacias:
                        QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
                    elif lineas_totales > 14:
                        QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
                    elif correcto == False:
                        QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
                    else:
                        formato = {
                            "cantidades" : cantidades,
                            "descripciones" : descripciones,
                            "valores_neto": valores_neto,
                            "creado_por" : self.datos_usuario[8]
                        }
                        detalle = json.dumps(formato)
                        #print('---------------------------------------------------')
                        try:
                           
                            if self.r_despacho.isChecked():
                                despacho = 'SI'
                            if self.seleccion == 'DIMENSIONADO':
                                if self.r_enchape.isChecked():
                                    enchapado = 'SI'
                                datos = ( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado, contacto,oce)
                                self.actualizar_pdf(datos,'dimensionado')

                                if self.conexion.root.actualizar_orden_dim( self.manual,self.interno,fecha_venta,self.tipo_doc,self.nro_doc,self.vendedor, self.nro_orden,nombre,telefono,str(fecha),detalle,despacho,enchapado,contacto,oce):
                                    QMessageBox.about(self,'EXITO','La orden de dimensionado fue ACTUALIZADA correctamente')
                                else:
                                    QMessageBox.about(self,'ERROR','La orden de dimensionado NO se actualizo, porque no se modificaron los datos que ya existian.')

                            elif self.seleccion == 'ELABORACION':
                                datos = ( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado, contacto,oce)
                                self.actualizar_pdf(datos,'elaboracion')

                                if self.conexion.root.actualizar_orden_elab( self.manual,self.interno,fecha_venta,self.tipo_doc,self.nro_doc,self.vendedor,  nombre,telefono, str(fecha), detalle, contacto, oce,despacho, self.nro_orden ):
                                    QMessageBox.about(self,'EXITO','La orden de elaboracion fue ACTUALIZADA correctamente')
                                else:
                                    QMessageBox.about(self,'ERROR','La orden de elaboracion NO se actualizo, porque no se modificaron los datos que ya existian.')
                            elif self.seleccion == 'CARPINTERIA':
                                datos = ( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado, contacto,oce)
                                self.actualizar_pdf(datos,'carpinteria')

                                if self.conexion.root.actualizar_orden_carp(self.manual,self.interno,fecha_venta,self.tipo_doc,self.nro_doc,self.vendedor, nombre,telefono, str(fecha), detalle, contacto, oce,despacho, self.nro_orden ):
                                    QMessageBox.about(self,'EXITO','La orden de carpinteria fue ACTUALIZADA correctamente')
                                else:
                                    QMessageBox.about(self,'ERROR','La orden de carpinteria NO se actualizo, porque no se modificaron los datos que ya existian.')
                            elif self.seleccion == 'PALLETS':
                                datos = ( str(self.nro_orden) , str(self.fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones,enchapado, contacto,oce)
                                self.actualizar_pdf(datos,'pallets')

                                if self.conexion.root.actualizar_orden_pall(self.manual,self.interno,fecha_venta,self.tipo_doc,self.nro_doc,self.vendedor, nombre,telefono, str(fecha), detalle, contacto, oce,despacho, self.nro_orden ):
                                    QMessageBox.about(self,'EXITO','La orden de pallets fue ACTUALIZADA correctamente')
                                else:
                                    QMessageBox.about(self,'ERROR','La orden de pallets NO se actualizo, porque no se modificaron los datos que ya existian.')
    
                        except EOFError:
                            QMessageBox.about(self,'Conexion' ,'Se perdio la conexion con el servidor')
                        except PermissionError:
                            QMessageBox.about(self,'ERROR' ,'La orden no se actualizo debido a que otro programa esta haciendo uso del archivo. Cierre dicho programa para continuar.')
                        #except AttributeError:
                         #   QMessageBox.about(self,'IMPORTANTE', 'Este mensaje se debe a que hubo un error al ingresar los datos a la base de datos. Contacte con el soporte')


                except ValueError:
                    if self.manual:
                        QMessageBox.about(self, 'ERROR', 'Solo ingrese numeros en el campo "Telefono" , "Numero Interno" y "Numero Documento" ') 
                    else:
                        QMessageBox.about(self, 'ERROR', 'Solo ingrese numeros en el campo "Telefono" ')          
            else:
                QMessageBox.about(self, 'Sugerencia', 'Ingrese un telefono antes de continuar')           
        else:
            QMessageBox.about(self, 'Sugerencia', 'Ingrese un nombre antes de continuar')

    def actualizar_pdf(self, lista,tipo):
        ruta = ( self.carpeta +'/ordenes/' + tipo +'_' + lista[0] + '.pdf' )  #NRO DE ORDEN  
        formato = self.carpeta +"/formatos/" + tipo +".jpg"
        agua = self.carpeta + "/formatos/despacho.png"
        hojas = 2
        if tipo == 'carpinteria':
            hojas = 1

        documento = canvas.Canvas(ruta)
        for pagina in range(hojas): #cantidad de hojas
            documento.setPageSize(( 216 * mm , 279 * mm))
            documento.drawImage( formato, 0* mm , 2 * mm , 216 *mm ,279 *mm )
            if self.r_despacho.isChecked():
                documento.setFillAlpha(0.6)
                documento.drawImage( agua , 83* mm , 30* mm , 100*mm ,100*mm , mask= 'auto')
                documento.drawImage( agua , 83* mm , (30+136)* mm , 100*mm ,100*mm , mask= 'auto')

            documento.setFillAlpha(1)
            documento.drawString( 0 * mm, 139.5 * mm ,'------------------------------------------------------------------------------------------')
            documento.drawString( 105 * mm, 139.5 * mm ,'----------------------------------------------------------------------------------------------')

            documento.rotate(90)

            documento.setFont('Helvetica',10)

            k = 2.5 #constante
            salto = 0
            for i in range(2):
                documento.setFont('Helvetica',9)
                documento.drawString( (28 + k + salto) *mm , -59.5 * mm , lista[2] )  #NOMBRE

                documento.drawString( (100 + k + salto) *mm , -66 * mm , str(lista[3]) )   #TELEFONO
                documento.drawString( (106 + k +salto) *mm , -59.5 * mm , lista[1]  )    #FECHA DE ORDEN

                
                
                if self.tipo_doc == 'FACTURA':
                    documento.drawString( (45 + k + salto) *mm , -85 * mm , str(self.nro_doc) )     #NRO FACTURA
                elif self.tipo_doc == 'BOLETA':
                    documento.drawString( (15 + k + salto) *mm , -85 * mm ,  str(self.nro_doc) )      #NRO BOLETA

                if tipo == 'dimensionado':
                    documento.drawString( (88 + k + salto) *mm , -94 * mm ,  lista[7] )   #ENCHAPE

                documento.drawString( (110 + k + salto) *mm , -85 * mm ,   lista[9]  )       #ORDEN DE COMPRA

                documento.drawString( (32 + k + salto) *mm , -66 * mm ,  lista[8] )   #CONTACTO
                if self.vendedor:
                    documento.drawString( (33+ salto) *mm , -205 * mm , self.vendedor ) #NOMBRE VENDEDOR
                
                documento.setFont('Helvetica-Bold',12)
                documento.drawString( (106 + k + salto) *mm , -44.5 * mm , lista[4]  ) #FECHA ESTIMADA
                documento.drawString( (106 + k + salto) *mm , -20.5 * mm , lista[0]  ) #NRO DE ORDEN

                documento.setFillColor(white)
                documento.setStrokeColor(white)
                documento.rect((5+ k + salto )* mm, -200 * mm ,80*mm,3*mm, fill=1 )
                documento.setFillColor(black)
                documento.setStrokeColor(black)
                documento.setFont('Helvetica-Bold',10)

                if pagina == 0:
                    if k == 2.5:
                        documento.drawString( (10+ salto) *mm , -200 * mm , 'ORIGINAL' )
                    else:
                        documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA ' + tipo.upper() )
                else:
                    if k == 2.5:
                        documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA CLIENTE' )
                    else:
                        documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA BODEGA' )
                documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )        
                salto += 139.5
                k = 0 

            #items
            cons = 108
            documento.setFont('Helvetica',9)
            align_cant_1 = 12
            align_descr_1 = 22
            align_cant_2 = 151
            align_descr_2 = 161

            cantidades = lista[5]
            descripciones = lista[6]
            i = 0
            while i < len(cantidades):
                documento.drawString(align_cant_1 *mm , -cons* mm , str(cantidades[i]) )  #cantidad
                
                documento.drawString(align_cant_2 *mm , -cons * mm , str(cantidades[i]) )

                cadenas = self.separar(descripciones[i])

                for cadena in cadenas:
                    documento.drawString(align_descr_1 *mm , -cons* mm , cadena )  #descripcion

                    documento.drawString(align_descr_2 *mm , -cons* mm , cadena ) #54 LONG MAXIMA
                    cons += 5
                i +=1

            documento.showPage()
        #documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )
        documento.save()
        sleep(1)
    
    def separar(self,cadena):
        lista = []
        iter = len(cadena)/54
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
        #print(cadena)
        #print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> 54:
        
            #print('long > 54:')
            aux = cadena[0:54]
            index = aux[::-1].find(' ')
        
            aux = aux[:(54-index)]
            #print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[54 - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                print('item vacio')
                #print('----------------------------------------------------')
            else:
                #print('LONG < 54: ' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista

    def ver_pdf(self):
        if self.seleccion == "DIMENSIONADO":
            abrir = self.carpeta + '/ordenes/' + 'dimensionado_' +str(self.nro_orden) + '.pdf'
        elif self.seleccion == "ELABORACION":
            abrir = self.carpeta + '/ordenes/'  + 'elaboracion_' +str(self.nro_orden) + '.pdf'
        elif self.seleccion == "CARPINTERIA":
            abrir = self.carpeta + '/ordenes/'  + 'carpinteria_' +str(self.nro_orden) + '.pdf'
        elif self.seleccion == "PALLETS":
            abrir = self.carpeta + '/ordenes/'  + 'pallets_' +str(self.nro_orden) + '.pdf'

        if os.path.isfile(abrir):
            subprocess.Popen([abrir], shell=True)
        else:
            QMessageBox.about(self, 'Error', 'No se encontro la orden de dimensionado en la carpeta "ordenes"')

    def anular(self):
        print(self.datos_usuario)
        if self.datos_usuario[4] == 'SI':
            dialog = InputDialog2('MOTIVO:' , 'INGRESE UN MOTIVO DE ANULACIÓN',self)
            dialog.resize(330,80)
            if dialog.exec():
                motivo  = dialog.getInputs()
                print(motivo)
                formato = {
                            "estado" : 'ANULADA',
                            "motivo" : motivo,
                            "usuario": self.datos_usuario[8]
                        }
                extra = json.dumps(formato)
                if self.seleccion == "DIMENSIONADO":
                    self.conexion.root.anular_orden('dimensionado', extra, self.nro_orden )
                elif self.seleccion == "ELABORACION":
                    self.conexion.root.anular_orden('elaboracion', extra, self.nro_orden )
                elif self.seleccion == "CARPINTERIA":
                    self.conexion.root.anular_orden('carpinteria', extra, self.nro_orden )
                elif self.seleccion == "PALLETS":
                    self.conexion.root.anular_orden('pallets', extra, self.nro_orden )
                QMessageBox.about(self,'EXITO' ,'Orden nro: ' + str(self.nro_orden) + ' Anulada')
                self.hide()
                self.parent().vaciar_tabla()
                self.parent().show()


        else:
            dialog = InputDialog('CLAVE:' ,'MOTIVO:', 'INGRESE UNA CLAVE Y MOTIVO DE ANULACION',self)
            dialog.resize(400,100)
            if dialog.exec():
                aux_clave , motivo = dialog.getInputs()
                claves = self.conexion.root.obtener_clave()
                clave = (aux_clave ,)
                if clave in claves:
                    print('clave valida')
                    formato = {
                            "estado" : 'ANULADA',
                            "motivo" : motivo,
                            "usuario": self.datos_usuario[0]
                        }
                    extra = json.dumps(formato)
                    if self.seleccion == "DIMENSIONADO":
                        self.conexion.root.anular_orden('dimensionado', extra, self.nro_orden )
                    elif self.seleccion == "ELABORACION":
                        self.conexion.root.anular_orden('elaboracion', extra, self.nro_orden )
                    elif self.seleccion == "CARPINTERIA":
                        self.conexion.root.anular_orden('carpinteria', extra, self.nro_orden )
                    elif self.seleccion == "PALLETS":
                        self.conexion.root.anular_orden('pallets', extra, self.nro_orden )
                    
                    self.conexion.root.eliminar_clave(aux_clave)
                    QMessageBox.about(self,'EXITO' ,'Orden nro: ' + str(self.nro_orden) + ' Anulada')
                    self.hide()
                    self.parent().vaciar_tabla()
                    self.parent().show()


                else:
                    print('clave invalida')
                    QMessageBox.about(self,'ERROR' ,'La clave ingresada no es valida')
                    
    def atras(self):
        self.parent().show()
        self.hide()

class Orden_manual(QMainWindow):
    def __init__(self, conn, clave, usuario , parent):
        super(Orden_manual, self).__init__(parent)
        uic.loadUi('dimensionado6.ui', self)
        self.conexion = conn
        self.tipo_doc = None
        self.nro_doc = None
        self.nro_orden = None
        self.nro_reingreso = None
        self.datos_usuario = usuario
        self.clave = clave
        #self.fecha_venta.setDate( datetime.now() )
        self.fecha.setDate( datetime.now().date() )
        self.fecha.setCalendarPopup(True)
        self.btn_registrar.clicked.connect(self.registrar)
        self.btn_atras.clicked.connect(self.atras)
        self.btn_atras_2.clicked.connect(self.atras2)
        self.btn_agregar.clicked.connect(self.agregar)
        self.btn_eliminar.clicked.connect(self.eliminar)
        self.btn_agregar_2.clicked.connect(self.agregar2)
        self.btn_eliminar_2.clicked.connect(self.eliminar2)
        self.btn_add.clicked.connect(self.add_descripcion)

        self.txt_descripcion_2.textChanged.connect(self.buscar_descripcion)
        self.txt_codigo.textChanged.connect(self.buscar_codigo)
        self.r_uso_interno.stateChanged.connect(self.cambiar_observacion)

        self.txt_descripcion_3.textChanged.connect(self.buscar_descripcion_2)
        self.btn_add_2.clicked.connect(self.add_descripcion_2)

        self.tableWidget.setColumnWidth(0,80)
        self.tableWidget.setColumnWidth(1,430)
        self.tableWidget.setColumnWidth(2,85)
        self.btn_registrar_2.clicked.connect(self.reingreso)
        self.tableWidget_2.setColumnWidth(1,80)
        self.tableWidget_2.setColumnWidth(0,430)
        self.tableWidget_2.setColumnWidth(2,85)
        
        
        self.carpeta = None
        self.inicializar()

    def inicializar(self):
        actual = os.path.abspath(os.getcwd())
        self.carpeta = actual.replace('\\' , '/')
        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_agregar.setIcon(QIcon('icono_imagen/agregar.ico'))
        self.btn_eliminar.setIcon(QIcon('icono_imagen/remover.ico'))
        self.btn_registrar.setIcon(QIcon('icono_imagen/guardar.ico'))
        
        self.btn_atras_2.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_agregar_2.setIcon(QIcon('icono_imagen/agregar.ico'))
        self.btn_eliminar_2.setIcon(QIcon('icono_imagen/remover.ico'))
        self.btn_registrar_2.setIcon(QIcon('icono_imagen/guardar.ico'))
    def agregar(self):
        if self.tableWidget.rowCount() <=16 :
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')

    def eliminar(self):
        fila = self.tableWidget.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tableWidget.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')

    def buscar_descripcion(self):
        self.productos.clear()
        descr = self.txt_descripcion_2.text()
        resultado = self.conexion.root.buscar_prod_descr(descr)
        for item in resultado:
            self.productos.addItem(item[1])
        #print(resultado)
    def buscar_codigo(self):
        self.productos.clear()
        codigo = self.txt_codigo.text()
        resultado = self.conexion.root.buscar_prod_cod(codigo)
        for item in resultado:
            self.productos.addItem(item[1])

    def add_descripcion(self):
        descripcion = self.productos.currentText()

        if self.tableWidget.rowCount() <=16 :
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
            self.tableWidget.setItem(fila, 0 , QTableWidgetItem( '0' ))
            self.tableWidget.setItem(fila, 1 , QTableWidgetItem( descripcion ))
            self.tableWidget.setItem(fila, 2 , QTableWidgetItem( '0' ))


        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')
    def cambiar_observacion(self):
        if self.r_uso_interno.isChecked():
            obs = 'uso carpinteria'
            self.txt_obs.clear() 
            self.txt_obs.appendPlainText(obs)
        else:
            self.txt_obs.clear() 

    def registrar(self):
        nombre = self.nombre.text()     #NOMBRE CLIENTE
        telefono = self.telefono.text() #TELEFONO
        fecha = self.fecha.date()  #FECHA ESTIMADA
        fecha = fecha.toPyDate()
        #self.tipo_doc = self.comboBox.currentText() #TIPO DE DOCUMENTO
        self.tipo_doc = None
        interno = 0 #NRO INTERNO
        self.nro_doc = None #NRO DOCUMENTO
        f_venta = None #FECHA DE  VENTA
        #f_venta = self.fecha_venta.dateTime() #FECHA DE  VENTA
        #f_venta = f_venta.toPyDateTime()
        #vendedor = self.txt_vendedor.text() #VENDEDOR
        vendedor = self.datos_usuario[8]
        observacion = self.txt_obs.toPlainText()
        lineas_totales = 0
        if nombre != '' and telefono != '' and observacion != '' :
            
            
            try:
                telefono = int(telefono)
                cant = self.tableWidget.rowCount()
                vacias = False
                correcto = True
                cantidades = []
                descripciones = []
                valores_neto = []
                i = 0
                while i< cant:
                    cantidad = self.tableWidget.item(i,0) #Collumna cantidades
                    descripcion = self.tableWidget.item(i,1) #Columna descripcion
                    neto = self.tableWidget.item(i,2) #Columna de valor neto
                    if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                        if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                            try: 
                                nueva_cant = cantidad.text().replace(',','.',3)
                                nuevo_neto = neto.text().replace(',','.',3)
                                cantidades.append( float(nueva_cant) )
                                descripciones.append(descripcion.text())

                                lineas = self.separar(descripcion.text())
                                lineas_totales = lineas_totales + len(lineas)

                                valores_neto.append(float(nuevo_neto))

                            except ValueError:
                                correcto = False
                        else:
                            vacias=True
                    else:
                        vacias = True
                    i+=1
                print('LINEAS TOTALES: ' + str(lineas_totales))
                if vacias:
                    QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
                elif lineas_totales > 14:
                    QMessageBox.about(self, 'Alerta' ,'Filas totales: '+str(lineas_totales) + ' - El maximo soportado por el formato de la orden es de 14 filas.' )
                elif correcto == False:
                    QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
                else:
                    formato = {
                        "cantidades" : cantidades,
                        "descripciones" : descripciones,
                        "valores_neto": valores_neto,
                        "creado_por" : self.datos_usuario[8]
                    }
                    detalle = json.dumps(formato)
                    try:
                        enchape = 'NO'
                        despacho = 'NO'                        
                        if self.r_despacho.isChecked():
                            despacho = 'SI'
                        
                        oce = self.oce.text()
                        fecha_orden = datetime.now().date()
                        cont = self.contacto.text()

                        if self.r_dim.isChecked():
                            if self.r_enchape.isChecked():
                                fecha = fecha + timedelta(days=2)
                                enchape = 'SI'
                            self.conexion.root.registrar_orden_dimensionado( interno , f_venta , nombre , telefono, str(fecha) , detalle, self.tipo_doc, self.nro_doc ,enchape,despacho,str(fecha_orden),cont,oce, vendedor )
                            
                            resultado = self.conexion.root.buscar_orden_dim_interno(interno)
                            if self.clave:
                                self.conexion.root.eliminar_clave(self.clave)
                            self.nro_orden = self.buscar_nro_orden(resultado)
                            self.conexion.root.actualizar_orden_dim_obser(observacion , self.nro_orden)

                            datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, enchape, cont,oce, vendedor)
                            self.crear_pdf(datos,'dimensionado')
                            boton = QMessageBox.question(self, 'Orden de dimensionado registrada correctamente', 'Desea abrir la Orden?')
                            if boton == QMessageBox.Yes:
                                self.ver_pdf('dimensionado')
                            self.hide()
                            self.parent().show()
                            
                        elif self.r_elab.isChecked():
                            self.conexion.root.registrar_orden_elaboracion( nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle,f_venta,vendedor)
                            if self.clave:
                                self.conexion.root.eliminar_clave(self.clave)
                            resultado = self.conexion.root.buscar_orden_elab_interno(interno)
                            self.nro_orden = self.buscar_nro_orden(resultado)
                            self.conexion.root.actualizar_orden_elab_obser(observacion , self.nro_orden)
                            datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, 'NO', cont, oce,vendedor)
                            self.crear_pdf(datos , 'elaboracion')
                            boton = QMessageBox.question(self, 'Orden de elaboracion registrada correctamente', 'Desea abrir la Orden?')
                            if boton == QMessageBox.Yes:
                                self.ver_pdf('elaboracion')
                            self.hide()
                            self.parent().show()
                        elif self.r_carp.isChecked():
                            self.conexion.root.registrar_orden_carpinteria( nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle, f_venta ,vendedor)
                            if self.clave:
                                self.conexion.root.eliminar_clave(self.clave)
                            resultado = self.conexion.root.buscar_orden_carp_interno(interno)
                            self.nro_orden = self.buscar_nro_orden(resultado)
                            self.conexion.root.actualizar_orden_carp_obser(observacion , self.nro_orden)
                            datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, 'NO', cont, oce, vendedor)
                            self.crear_pdf(datos , 'carpinteria')
                            boton = QMessageBox.question(self, 'Orden de elaboracion registrada correctamente', 'Desea abrir la Orden?')
                            if boton == QMessageBox.Yes:
                                self.ver_pdf('carpinteria')
                            self.hide()
                            self.parent().show()
                        elif self.r_pall.isChecked():
                            self.conexion.root.registrar_orden_pallets( nombre,telefono,str(fecha_orden), str(fecha),self.nro_doc,self.tipo_doc,cont,oce, despacho, interno ,detalle, f_venta,vendedor)
                            if self.clave:
                                self.conexion.root.eliminar_clave(self.clave)
                            resultado = self.conexion.root.buscar_orden_pall_interno(interno)
                            self.nro_orden = self.buscar_nro_orden(resultado)
                            self.conexion.root.actualizar_orden_pall_obser(observacion , self.nro_orden)
                            datos = ( str(self.nro_orden) , str(fecha_orden.strftime("%d-%m-%Y")), nombre , telefono, str(fecha.strftime("%d-%m-%Y")) , cantidades, descripciones, 'NO', cont, oce,vendedor)
                            self.crear_pdf(datos , 'pallets')
                            boton = QMessageBox.question(self, 'Orden de elaboracion registrada correctamente', 'Desea abrir la Orden?')
                            if boton == QMessageBox.Yes:
                                self.ver_pdf('pallets')
                            self.hide()
                            self.parent().show()
                        else:
                            QMessageBox.about(self, 'ERROR', 'Seleccione un tipo de orden a generar, antes de proceder a registrar')    

                    except EOFError:
                        QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')   
                    #except AttributeError:

                     #   QMessageBox.about(self,'IMPORTANTE', 'Este mensaje se debe a que hubo un error al ingresar los datos a la base de datos. Contacte con el soporte')


            except ValueError:
                QMessageBox.about(self, 'ERROR', 'Solo ingrese Numeros en los campos "Telefono", "Numero de documento" y "Numero interno" ')          
        else:
                QMessageBox.about(self, 'Sugerencia', 'Los campos "Nombre" , "Telefono" y "Observacion" son obligatorios')           


    def crear_pdf(self, lista,tipo):#ORDEN MANUAL

        ruta = ( self.carpeta +'/ordenes/' + tipo +'_' + lista[0] + '.pdf' )  #NRO DE ORDEN  
        formato = self.carpeta +"/formatos/" + tipo +".jpg"
        agua = self.carpeta + "/formatos/despacho.png"
        uso_interno = self.carpeta + "/formatos/uso interno.png"
        hojas = 2
        if tipo == 'carpinteria':
            hojas = 1
        try:
            documento = canvas.Canvas(ruta)
            for pagina in range(hojas): #cantidad de hojas
                documento.setPageSize(( 216 * mm , 279 * mm))
                documento.drawImage( formato, 0* mm , 2 * mm , 216 *mm ,279 *mm )
                if self.r_despacho.isChecked():
                    documento.setFillAlpha(0.6)
                    documento.drawImage( agua , 83* mm , 30* mm , 100*mm ,100*mm , mask= 'auto')
                    documento.drawImage( agua , 83* mm , (30+136)* mm , 100*mm ,100*mm , mask= 'auto')
                elif self.r_uso_interno.isChecked():
                    documento.setFillAlpha(0.6)
                    documento.drawImage( uso_interno , 100* mm , 30* mm , 69*mm ,94.5 *mm , mask= 'auto')
                    documento.drawImage( uso_interno , 100* mm , (30+136)* mm , 69*mm ,94.5*mm , mask= 'auto')
                
                documento.setFillAlpha(1)
                documento.drawString( 0 * mm, 139.5 * mm ,'------------------------------------------------------------------------------------------')
                documento.drawString( 105 * mm, 139.5 * mm ,'----------------------------------------------------------------------------------------------')

                documento.rotate(90)

                documento.setFont('Helvetica',10)

                k = 2.5 #constante
                salto = 0
                for i in range(2):

                    if self.r_facturar.isChecked():
                        documento.setFont('Helvetica',11)
                        documento.setFillAlpha(0.6)
                        documento.drawString( (53 + k + salto) *mm , -20.5 * mm , '"POR FACTURAR"' )  #por facturar
                        documento.setFillAlpha(1)

                    documento.setFont('Helvetica',9)
                    documento.drawString( (28 + k + salto) *mm , -59.5 * mm , lista[2] )  #NOMBRE

                    documento.drawString( (100 + k + salto) *mm , -66 * mm , str(lista[3]) )   #TELEFONO
                    documento.drawString( (106 + k +salto) *mm , -59.5 * mm , lista[1]  )    #FECHA DE ORDEN

                    
                    
                    '''if self.tipo_doc == 'FACTURA':
                        documento.drawString( (45 + k + salto) *mm , -85 * mm , str(self.nro_doc) )     #NRO FACTURA
                    elif self.tipo_doc == 'BOLETA':
                        documento.drawString( (15 + k + salto) *mm , -85 * mm ,  str(self.nro_doc) )      #NRO BOLETA '''

                    if tipo == 'dimensionado':
                        documento.drawString( (88 + k + salto) *mm , -94 * mm ,  lista[7] )   #ENCHAPE

                    documento.drawString( (110 + k + salto) *mm , -85 * mm ,   lista[9]  )       #ORDEN DE COMPRA

                    documento.drawString( (32 + k + salto) *mm , -66 * mm ,  lista[8] )   #CONTACTO

                    if lista[10]:
                        documento.drawString( (33+ salto) *mm , -205 * mm , lista[10] ) #NOMBRE VENDEDOR
                    
                    documento.setFont('Helvetica-Bold',12)
                    documento.drawString( (106 + k + salto) *mm , -44.5 * mm , lista[4]  ) #FECHA ESTIMADA
                    documento.drawString( (106 + k + salto) *mm , -20.5 * mm , lista[0]  ) #NRO DE ORDEN

                    documento.setFillColor(white)
                    documento.setStrokeColor(white)
                    documento.rect((5+ k + salto )* mm, -200 * mm ,80*mm,3*mm, fill=1 )
                    documento.setFillColor(black)
                    documento.setStrokeColor(black)
                    documento.setFont('Helvetica-Bold',10)

                    if pagina == 0:
                        if k == 2.5:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'ORIGINAL' )
                        else:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA ' + tipo.upper() )
                    else:
                        if k == 2.5:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA CLIENTE' )
                        else:
                            documento.drawString( (10+ salto) *mm , -200 * mm , 'COPIA BODEGA' )
                    documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )        
                    salto += 139.5
                    k = 0 

                #items
                cons = 108
                documento.setFont('Helvetica',9)
                align_cant_1 = 12
                align_descr_1 = 22
                align_cant_2 = 151
                align_descr_2 = 161

                cantidades = lista[5]
                descripciones = lista[6]
                i = 0
                while i < len(cantidades):
                    documento.drawString(align_cant_1 *mm , -cons* mm , str(cantidades[i]) )  #cantidad
                    
                    documento.drawString(align_cant_2 *mm , -cons * mm , str(cantidades[i]) )

                    cadenas = self.separar(descripciones[i])

                    for cadena in cadenas:
                        documento.drawString(align_descr_1 *mm , -cons* mm , cadena )  #descripcion

                        documento.drawString(align_descr_2 *mm , -cons* mm , cadena ) #54 LONG MAXIMA
                        cons += 5
                    i +=1

                documento.showPage()
            #documento.drawString( (10+ salto) *mm , -205 * mm , 'VENDEDOR:' )
            documento.save()
            sleep(1)
        except PermissionError:
            QMessageBox.about(self,'ERROR', 'Otro programa esta modificando este archivo por lo cual no se puede modificar actualmente.')
    
    def separar(self,cadena):
        lista = []
        iter = len(cadena)/54
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
        #print(cadena)
        #print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> 54:
        
            #print('long > 54:')
            aux = cadena[0:54]
            index = aux[::-1].find(' ')
        
            aux = aux[:(54-index)]
            #print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[54 - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                print('item o cadena vacio')
                #3print('----------------------------------------------------')
            else:
                #print('LONG < 54: ' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista
    
    def ver_pdf(self, tipo):    
        abrir = self.carpeta + '/ordenes/' +tipo +'_' +str(self.nro_orden) + '.pdf'
        subprocess.Popen([abrir], shell=True)

    def buscar_nro_orden(self,tupla): #BUSCA EL NRO DE ORDEN CREADO RECIENTEMENTE, POR ENDE EL MAYOR.
        mayor = 0
        for item in tupla:
            if item[0] > mayor :
                mayor = item[0]
        return mayor

    def generar_pdf(self,datos): #REINGRESO MANUAL
        
        documento = canvas.Canvas(self.carpeta +'/reingresos/reingreso_' + str(datos[0]) + '.pdf')
        imagen =  self.carpeta + "/formatos/reingreso_solo.jpg" 

        

        documento.setPageSize(( 216 * mm , 279 * mm))
        documento.drawImage( imagen, 0* mm , 0 * mm , 216 *mm , 139.5 *mm )
        documento.drawImage( imagen, 0* mm , 139.5 * mm , 216 *mm , 139.5 *mm )
        documento.drawString( 0*mm , 139.5 *mm , '------------------------------------------------------------------------------')
        documento.drawString( 108*mm , 139.5 *mm , '----------------------------------------------------------------------------')
        salto = 0
      
        for i in range(2):
            documento.setFont('Helvetica',9)
            
            if datos[2] == 'FACTURA':
                documento.drawString(129*mm, (106.5 + salto )*mm , str(datos[3])  )   #NUMERO DOCUMENTO , FACTURA
            elif datos[2] == 'BOLETA':
                documento.drawString(52* mm, (106.5+ salto )*mm , str(datos[3])  )   #NUMERO DOCUMENTO , BOLETA
           
            lista = self.separar2(datos[5],94) #DESCRIPCION
            
            print(len( datos[5] ))
            k = 0 
            j = 0 
            for item in lista:
                documento.drawString(20* mm, (92.5 + salto - k)*mm , lista[j] )  #descripcion del problema
                #documento.drawString(20* mm, (92.5 + salto - k )*mm , descr )  #descripcion del problema
                k += 6
                j += 1
            
            lista = self.separar2( datos[7] , 85) #SOLUCION
            print(len(datos[7]))
            k = 0 
            j = 0
            for item in lista:
                documento.drawString(40*mm, (39 + salto - k )*mm , lista[j] )  #solucion del problema
                k += 6
                j += 1
            
            cants = datos[8]
            descrs = datos[9]
            netos = datos[10]
            p = 0
            q = 0
            for item in cants:
                documento.drawString(150*mm, (65 + salto - q )*mm , str(cants[p]) )  #cantidad
                documento.drawString(170*mm, (65 + salto - q )*mm , str(netos[p]) )  #neto
                cadenas = self.separar2(descrs[p] , 60 ) 
                for cadena in cadenas:
                    documento.drawString(20*mm, (65 + salto -q )*mm , cadena)  #descripcion
                    q += 5
                p +=1

            documento.setFont('Helvetica-Bold', 9 )

            documento.drawString(177*mm, (124.5 + salto )*mm ,  str(datos[0]) )  #NRO DE REINGRESO
            documento.drawString(177*mm, (115.5 + salto ) *mm , datos[1] )    #FECHA DEL REINGRESO

            if datos[6] == 'DIMENSIONADO':
                documento.drawString(73*mm, (77 + salto )*mm , 'X' )  #PROCESO DIMENSIONADO
            elif datos[6] == 'ELABORACION':    
                documento.drawString(150*mm, (77 + salto )*mm , 'X' )  #PROCESO ELABORACION
            elif datos[6] == 'CARPINTERIA':
                documento.drawString(111*mm, (77 + salto )*mm , 'X' )  #PROCESO CARPINTERIA
            elif datos[6] == 'PALLETS':
                documento.drawString(178.5*mm, (77 + salto )*mm , 'X' )  #PROCESO PALLETS


            if datos[4] == 'CAMBIO':
                documento.drawString(53*mm, (99.5 + salto )*mm , 'X' )  #motivo cambio
            elif datos[4] == 'DEVOLUCION':
                documento.drawString(92*mm, (99.5 + salto )*mm , 'X' )   #motivo devolucion
            else:
                documento.setFont('Helvetica',9)
                documento.drawString(128*mm, (99.5 + salto )*mm , datos[4] )  #motivo otro

            salto += 139.5 
        documento.save()

    def atras(self):
        self.hide()
        self.parent().show()
    
    def reingreso(self):
        nro_orden = self.txt_orden.text()           #NUMERO DE ORDEN
        tipo_doc = self.comboBox_2.currentText() #TIPO DE DOCUMENTO
        nro_doc = self.txt_nro_doc.text()        #NUMERO DE DOCUMENTO
        fecha = datetime.now().date()              #FECHA DE REINGRESO
        motivo = ''                                 #MOTIVO
        if self.r_cambio.isChecked():
            motivo = 'CAMBIO'
        elif self.r_devolucion.isChecked():
            motivo = 'DEVOLUCION'
        elif self.r_otro.isChecked():
            motivo = self.txt_otro.text()
        
        proceso = None
        if self.r_d.isChecked():
            proceso = 'DIMENSIONADO'
        elif self.r_e.isChecked():
            proceso = 'ELABORACION'
        elif self.r_c.isChecked():
            proceso = 'CARPINTERIA'
        elif self.r_p.isChecked():
            proceso = 'PALLETS'
        descr = self.txt_descripcion.toPlainText()   #descripcion
        solucion = self.txt_solucion.toPlainText()   #solucion
        lineas = 0
        if motivo != '' and descr != '' and solucion != '' :
            cant = self.tableWidget_2.rowCount()
            vacias = False #Determinna si existen campos vacios
            correcto = True #Determina si los datos estan escritos correctamente. campos cantidad y valor son numeros.
            cantidades = []
            descripciones = []
            valores_neto = []
            i = 0
            while i< cant:
                descripcion = self.tableWidget_2.item(i,0) #Collumna descripcion
                cantidad = self.tableWidget_2.item(i,1) #Columna cantidad
                neto = self.tableWidget_2.item(i,2) #Columna de valor neto
                if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                    if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                        try: 
                            nueva_cant = cantidad.text().replace(',','.',3)
                            nuevo_neto = neto.text().replace(',','.',3)
                            cantidades.append( float(nueva_cant) )
                            descripciones.append(descripcion.text())
                            linea = self.separar2(descripcion.text(), 60 )
                            lineas += len(linea)
                            print('total de lineas: '+ str(lineas))
                            valores_neto.append(float(nuevo_neto))

                        except ValueError:
                            correcto = False
                    else:
                        vacias=True
                else:
                    vacias = True
                i+=1
            if vacias:
                QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
            elif correcto == False:
                QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
            elif lineas > 4:
                QMessageBox.about(self, 'Alerta' ,'El maximo de filas por el formato de impresion es de 4.' )
            else:
                #print(cantidades)
                #print(valores_neto)
                formato = {
                        "cantidades" : cantidades,
                        "descripciones" : descripciones,
                        "valores_neto": valores_neto,
                        "creado_por" : self.datos_usuario[8]
                    }
                detalle = json.dumps(formato)
                try:
                    nro_orden = int(nro_orden)
                    nro_doc = int(nro_doc)

                    if self.conexion.root.registrar_reingreso( str(fecha), tipo_doc, nro_doc, nro_orden, motivo, descr, proceso, detalle,solucion):
                        resultado = self.conexion.root.obtener_max_reingreso()
                        self.nro_reingreso = resultado[0]
                        print('max nro reingreso: ' + str(resultado[0]) + ' de tipo: ' + str(type(resultado[0])))
                        datos = (resultado[0], str(fecha) , tipo_doc , nro_doc , motivo , descr , proceso , solucion, cantidades, descripciones, valores_neto)
                        self.generar_pdf(datos)
                        if self.clave:
                                self.conexion.root.eliminar_clave(self.clave)

                        boton = QMessageBox.question(self, 'Reingreso registrado correctamente', 'Desea ver el reingreso?')
                        if boton == QMessageBox.Yes:
                            self.abrir_pdf()

                        self.hide()
                        self.parent().show()
                    else:
                        QMessageBox.about(self,'ERROR','404 NOT FOUND. Contacte con Don Huber ...problemas al registrar')
                except ValueError:
                    QMessageBox.about(self,'ERROR','Ingresar solo numeros en los campos: "NUMERO DE ORDEN" y "NUMERO DE DOCUMENTO" ')


        else:
            QMessageBox.about(self,'Datos incompletos','Los campos "descripcion" , "solucion" son obligatiorios. Como tambien si selecciona "OTROS" debe rellenar su campo')

    def abrir_pdf(self):
        ruta = self.carpeta + '/reingresos/reingreso_' + str(self.nro_reingreso) + '.pdf'
        subprocess.Popen([ruta], shell=True)

    def separar2(self, cadena, long):
        lista = []
        iter = len(cadena)/long
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
        print(cadena)
        print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> long:
        
            #print('long > '+ str(long) +':')
            aux = cadena[0:long]
            index = aux[::-1].find(' ')
        
            aux = aux[:(long-index)]
            #print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[long - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                print('item vacio')
                #print('----------------------------------------------------')
            else:
                #print('fin long < '+ str(long) +':' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista

    def agregar2(self):
        if self.tableWidget_2.rowCount() <=16 :
            fila = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')

    def eliminar2(self):
        fila = self.tableWidget_2.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tableWidget_2.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    def atras2(self):
        self.hide()
        self.parent().show()
    def buscar_descripcion_2(self):
        self.productos_2.clear()
        descr = self.txt_descripcion_3.text()
        resultado = self.conexion.root.buscar_prod_descr(descr)
        for item in resultado:
            self.productos_2.addItem(item[1])
    def add_descripcion_2(self):
        descripcion = self.productos_2.currentText()

        if self.tableWidget_2.rowCount() < 4 :
            fila = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(fila)
            self.tableWidget_2.setItem(fila, 0 , QTableWidgetItem( descripcion ))
            self.tableWidget_2.setItem(fila, 1 , QTableWidgetItem( '0' ))
            self.tableWidget_2.setItem(fila, 2 , QTableWidgetItem( '0' ))


        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas soportado para la impresión de un reingreso. Intente hacer otro reingreso para los items faltantes ')

class Informes(QMainWindow):
    def __init__(self, conn, parent):
        super(Informes, self).__init__(parent)
        uic.loadUi('informes.ui', self)
        self.conexion = conn
        self.dir_formatos = None
        self.dir_informes = None
        self.lista_informes = None
        self.groupBox.hide()
        self.inicializar()
        self.tableWidget.setColumnWidth(0,371)
        self.btn_atras.clicked.connect(self.atras)
        self.r_orden.setChecked(True)
        self.d_inicio.setDate( datetime.now().date() )
        self.d_inicio.setCalendarPopup(True)
        self.d_termino.setDate( datetime.now().date() )
        self.d_termino.setCalendarPopup(True)
        self.btn_generar.clicked.connect(self.generar)
        self.btn_actualizar.clicked.connect(self.actualizar)
        self.btn_abrir.clicked.connect(self.abrir)
        self.btn_eliminar.clicked.connect(self.eliminar)
        self.comboBox.currentIndexChanged['QString'].connect(self.vista_reingreso)
        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))
        self.btn_generar.setIcon(QIcon('icono_imagen/continuar.ico'))
        self.btn_actualizar.setIcon(QIcon('icono_imagen/actualizar.ico'))
        self.btn_abrir.setIcon(QIcon('icono_imagen/continuar.ico'))
        self.btn_eliminar.setIcon(QIcon('icono_imagen/borrar.ico'))

    def inicializar(self):
        actual = os.path.abspath(os.getcwd())
        actual = actual.replace('\\' , '/')

        self.dir_formatos = actual + '/formatos/'
        self.dir_informes = actual + '/informes/'

        informes = os.listdir(self.dir_informes)
        for item in informes:
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( item ) )  #nombre informe
        

    def actualizar(self):
        self.tableWidget.setRowCount(0)
        informes = os.listdir(self.dir_informes)
        for item in informes:
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
            self.tableWidget.setItem(fila , 0 , QTableWidgetItem( item ) )  #nombre informe
    def abrir(self):
        seleccion = self.tableWidget.selectedItems()
        if seleccion != [] :
            nombre = seleccion[0].text()
            abrir = self.dir_informes + nombre
            if os.path.isfile(abrir):
                subprocess.Popen([abrir], shell=True)
            else:
                QMessageBox.about(self,'ERROR', 'El archivo no se encontro en la carpeta "informes" ')
        else:
            QMessageBox.about(self,'Sugerencia', 'Primero seleccione el nombre del informe para poder abrirlo ')

    def eliminar(self):
        seleccion = self.tableWidget.selectedItems()
        if seleccion != [] :
            nombre = seleccion[0].text()
            ruta = self.dir_informes + nombre
            if os.path.isfile(ruta):
                try:
                    os.remove(ruta)
                    self.actualizar()
                except PermissionError:
                    QMessageBox.about(self,'ERROR', 'No se pudo eliminar el informe, ya que archivo esta siendo utilizado por otro programa')
            else:
                QMessageBox.about(self,'ERROR', 'El archivo no se encontro')
        else:
            QMessageBox.about(self,'Sugerencia', 'Primero seleccione el nombre del informe para poder abrirlo ')

    def generar(self):
        tipo_orden = self.comboBox.currentText()
        inicio = self.d_inicio.date()
        inicio = inicio.toPyDate()
        termino = self.d_termino.date()
        termino = termino.toPyDate()

        nombre =self.dir_informes + tipo_orden + '_'+ str( inicio.strftime("%d-%m-%Y")) + '_HASTA_' + str( termino.strftime("%d-%m-%Y")) + '.xlsx' 
        datos = None
        if self.r_orden.isChecked():
            try:
                if tipo_orden == 'DIMENSIONADO':
                    datos = self.conexion.root.informe_dimensionado(str(inicio) , str(termino))
                    self.informe_dimensionado(datos,nombre)
                    
                elif tipo_orden == 'ELABORACION':
                    datos = self.conexion.root.informe_elaboracion(str(inicio) , str(termino))
                    self.informe_generico(datos,nombre)
                elif tipo_orden == 'CARPINTERIA':
                    datos = self.conexion.root.informe_carpinteria(str(inicio) , str(termino))
                    self.informe_generico(datos,nombre)
                elif tipo_orden == 'PALLETS':
                    datos = self.conexion.root.informe_pallets(str(inicio) , str(termino))
                    self.informe_generico(datos,nombre)
                elif tipo_orden == 'REINGRESO':
                    if self.r_d.isChecked() or self.r_e.isChecked() or self.r_c.isChecked() or self.r_p.isChecked() :
                        acept = []
                        nombre =self.dir_informes + tipo_orden
                        entremedio = ''
                        if self.r_d.isChecked():
                            acept.append('DIMENSIONADO')
                            entremedio = entremedio + '_D'
                        if self.r_e.isChecked():
                            acept.append('ELABORACION')
                            entremedio = entremedio + '_E'
                        if self.r_c.isChecked():
                            acept.append('CARPINTERIA')
                            entremedio = entremedio + '_C'
                        if self.r_p.isChecked():
                            acept.append('PALLETS')
                            entremedio = entremedio + '_P'
                        fin = '_'+ str( inicio.strftime("%d-%m-%Y")) + '_HASTA_' + str( termino.strftime("%d-%m-%Y")) + '.xlsx' 
                        nombre = nombre + entremedio + fin
                        datos = self.conexion.root.informe_reingreso(str(inicio), str(termino))
                        self.informe_reingreso(datos, acept, nombre)
                    else:
                        QMessageBox.about(self,'CONSEJO', 'SELECCIONE ALMENOS UN TIPO DE PROCESO PARA CCREAR EL INFORME DE REINGRESO')

                if datos:
                    self.actualizar()
                    QMessageBox.about(self,'EXITO', 'Informe generado correctamente')
                
            except PermissionError:
                QMessageBox.about(self,'ERROR',  'Otro programa tiene abierto el documento EXCEL. Intente cerrandolo y luego proceda a generar el EXCEL')
            except EOFError:
                QMessageBox.about(self,'ERROR','No hubo respuesta desde el servidor')

        elif self.r_venta.isChecked():
            #EN PROCESO ALGUN DIA 
            x = 0
        

    def informe_dimensionado(self,datos,nombre):
        if datos:
           
            wb = Workbook()
            ws = wb.active
            encabezado = ['TIPO DOCUMENTO','NRO DOCUMENTO','NOMBRE CLIENTE', 'NRO ORDEN', 'TELEFONO','DESCRIPCION','CANTIDAD','PRECIO NETO', 'DIMENSIONADOR','VENDEDOR','FECHA VENTA','FECHA ORDEN','FECHA INGRESO','FECHA ESTIMADA','FECHA REAL','ENCHAPADO','DESPACHO','CONTACTO','ORDEN COMPRA','OBSERVACION','ESTADO','MOTIVO','CREADO POR']

            ws.append(encabezado)
            f_ven = None

            for item in datos:
                if item[8]:
                    fecha_venta = datetime.fromisoformat( str( item[8] )) 
                    f_ven =  str( fecha_venta.strftime( "%d-%m-%Y %H:%M:%S" ) )  #FECHA DE VENTA
                f_ing = 'No asignada'
                f_real = 'No asignada'
                if item[4]:
                    ingreso = datetime.fromisoformat(str( item[4]) )
                    f_ing =  str(ingreso.strftime("%d-%m-%Y") ) #FECHA DE INGRESO
                estimada = datetime.fromisoformat(str( item[3]) )
                f_est =  str(estimada.strftime("%d-%m-%Y") )  #FECHA ESTIMADA DE ENTREGA
                if item[5]:
                    real = datetime.fromisoformat(str( item[5]) )
                    f_real =  str(real.strftime("%d-%m-%Y") ) #FECHA REAL DE ENTREGA
                fecha_orden = datetime.fromisoformat(str( item[14]) )
                f_ord =  str(fecha_orden.strftime("%d-%m-%Y") )  #FECHA DE ORDEN    
                detalle = json.loads( item[6])
                cant = detalle['cantidades']
                desc = detalle['descripciones']
                net = detalle["valores_neto"]
                try:
                    creador = detalle["creado_por"]
                except KeyError:
                    creador = 'No registrado'

                if item[19]:
                    extra = json.loads( item[19] )
                    estado = extra["estado"]
                    motivo = extra["motivo"]
                    
                else:
                    estado = 'VALIDA'
                    motivo = 'Ninguno'
                j = 0
                while j < len(cant):
                    fila = [item[11],item[10],item[1], item[0], item[2], desc[j],cant[j],net[j] ,item[9],item[17],f_ven,f_ord,f_ing,f_est,f_real,item[12],item[13],item[15],item[16],item[18],estado,motivo,creador]
                    ws.append(fila)
                    j +=1
            filas_total = ws.max_row
            tab = Table(displayName="tabla_dimensionado" , ref="A1:W"+ str(filas_total) )
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
            dim_holder['C'] = ColumnDimension(ws, min=3 ,max = 3, width=30)
            dim_holder['F'] = ColumnDimension(ws, min=6 ,max = 6, width=35)
            dim_holder['J'] = ColumnDimension(ws, min=10 ,max = 10, width=30)
            ws.column_dimensions = dim_holder
            ws.add_table(tab)
            wb.save(nombre)
        
        else:
            QMessageBox.about(self, 'ERROR', 'No se encontraron ordenes de trabajo para el rango de fechas definido')

    def informe_generico(self,datos,nombre):
        if datos:
            wb = Workbook()
            ws = wb.active
            encabezado = ['TIPO DOCUMENTO','NRO DOCUMENTO','NOMBRE CLIENTE', 'NRO ORDEN', 'TELEFONO','DESCRIPCION','CANTIDAD','PRECIO NETO','VENDEDOR','FECHA VENTA','FECHA ORDEN','FECHA ESTIMADA','FECHA REAL','DESPACHO','CONTACTO','ORDEN COMPRA','OBSERVACION','ESTADO','MOTIVO','CREADA POR']
            ws.append(encabezado)
            fv = None
            for item in datos:
                td =  item[7] #TIPO DOCUMENTO
                nd =  item[6]  #NUMERO DOCUMENTO  
                cl =  item[1]  #CLIENTE
                no =  item[0]  #NUMERO DE ORDEN
                tel =  item[2]  #TELEFONO
                detalle = json.loads( item[12]  )
                cant = detalle['cantidades']  #cantidades
                desc = detalle['descripciones'] #descripciones
                net = detalle["valores_neto"]  #valores neto
                try:
                    creador = detalle["creado_por"]
                except KeyError:
                    creador = 'No registrado'

                if item[13]:
                    fecha_venta = datetime.fromisoformat( str( item[13] ))
                    fv =  str( fecha_venta.strftime( "%d-%m-%Y %H:%M:%S" ) )  #FECHA DE VENTA
                estimada = datetime.fromisoformat(str( item[4]) )
                fe =  str(estimada.strftime("%d-%m-%Y") )  #FECHA ESTIMADA DE ENTREGA
                fr = 'No asignada'
                if item[5]:
                    real = datetime.fromisoformat(str( item[5]) )
                    fr = str(real.strftime("%d-%m-%Y") )   #FECHA REAL
                de =  item[10]  #DESPACHO
                co =  item[8]  #CONTACTO
                oc =  item[9]  #ORDEN DE COMPRA
                ve =  item[14]  #VENDEDOR
                fecha_orden = datetime.fromisoformat(str( item[3]) )
                fo =  str(fecha_orden.strftime("%d-%m-%Y") )  #FECHA DE ORDEN
                ob =  item[15] #OBSERVACION
                if item[16]: #ESTADO
                    extra = json.loads( item[16] )
                    estado = extra["estado"]
                    motivo = extra["motivo"]
                    
                else:
                    estado = 'VALIDA'
                    motivo = 'Ninguno'
                j = 0
                while j < len(cant):
                    fila = [ td,nd,cl,no,tel,desc[j],cant[j],net[j],ve ,fv,fo,fe,fr,de,co,oc,ob,estado,motivo,creador]
                    ws.append(fila)
                    j +=1

            filas_total = ws.max_row
            tab = Table(displayName="tabla1" , ref="A1:T"+ str(filas_total) )
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
            dim_holder['C'] = ColumnDimension(ws, min=3 ,max = 3, width=30)
            dim_holder['F'] = ColumnDimension(ws, min=6 ,max = 6, width=35)
            dim_holder['I'] = ColumnDimension(ws, min=9 ,max = 9, width=30)
            ws.column_dimensions = dim_holder
            ws.add_table(tab)
            wb.save(nombre)
            
        else:
            QMessageBox.about(self, 'ERROR', 'No se encontraron ordenes de trabajo para el rango de fechas definido')

    def vista_reingreso(self):
        if self.comboBox.currentText() == 'REINGRESO':
            print('modo reingreso')
            self.groupBox.show()
            self.r_orden.hide()
            self.r_venta.hide()
            self.lb_buscar.show()
            self.lb_buscar.setText('FECHA DE CREACION')
        else:
            self.groupBox.hide()
            self.lb_buscar.hide()
            self.r_orden.show()
            self.r_venta.show()

    def informe_reingreso(self, datos,acept, nombre):
        if datos:
            wb = Workbook()
            ws = wb.active
            ws.title = 'REINGRESO_1'
            encabezado = ['NUMERO REINGRESO','FECHA REINGRESO','NUMERO DE ORDEN', 'TIPO DOCUMENTO', 'NUMERO DOCUMENTO', 'PROCESO','MERCADERIA','CANTIDAD','VALOR NETO','MOTIVO','DESCRIPCIÓN' ,'SOLUCIÓN','CREADO POR']
            ws.append(encabezado)
            for item in datos:
                nro_re =  str( item[0] ) #NRO REINGRESO INT
                reingreso = datetime.fromisoformat(str( item[1]) )
                fr =  str(reingreso.strftime("%d-%m-%Y") )  #FECHA REINGRESO
                td =  item[2] #TIPO DOCUMENTO   STR
                nd =  str( item[3] )  #NUMERO DOCUMENTO  INT
                no =  str( item[4] )  #NUMERO DE ORDEN    INT
                mot =  item[5]  #MOTIVO   STR
                descripcion =  item[6]  #DESCRIPCION  STR
                proc =  item[7]  #PROCESO   STR
                if proc in acept: #se filtra
                    detalle = json.loads( item[8]  ) #DETALLE
                    cant = detalle['cantidades']  #cantidades
                    merc = detalle['descripciones'] #mercaderia
                    net = detalle["valores_neto"]  #valores neto
                    try:
                        creador = detalle["creado_por"]
                    except KeyError:
                        creador = 'No registrado'

                    sol =  item[9] #SOLUCION  STR
                    j = 0
                    while j < len(cant):
                        fila = [ nro_re, fr, no, td, nd, proc, merc[j], cant[j], net[j], mot, descripcion ,sol, creador]
                        ws.append(fila)
                        j +=1

            filas_total = ws.max_row
            tab = Table(displayName="tabla1" , ref="A1:M"+ str(filas_total) )
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)
            dim_holder['F'] = ColumnDimension(ws, min=6 ,max = 6, width=16)  #PROCESO
            dim_holder['G'] = ColumnDimension(ws, min=7 ,max = 7, width=45) #MERCADERIA 45px
            dim_holder['I'] = ColumnDimension(ws, min=9 ,max = 9, width=14) #VALOR NETO 10px
            dim_holder['J'] = ColumnDimension(ws, min=10 ,max = 10, width=14) #MOTIVO
            dim_holder['K'] = ColumnDimension(ws, min=11 ,max = 11, width=30)  #DESCRIPCION
            dim_holder['L'] = ColumnDimension(ws, min=12 ,max = 12, width=30)  #SOLUCION
            ws.column_dimensions = dim_holder
            ws.add_table(tab)
            wb.save(nombre)
        else: 
            QMessageBox.about(self,'ERROR', 'No se encontraron datos de reingreso para las fechas ingresadas.')
    def atras(self):
        self.hide()
        self.parent().show()

class Reingreso(QMainWindow):
    def __init__(self, conn , orden , tipo ,parent):
        super( Reingreso, self).__init__(parent)
        uic.loadUi('re_ingreso.ui' , self)
        self.seleccion = tipo
        self.tableWidget.setColumnWidth(1,80)
        self.tableWidget.setColumnWidth(0,430)
        self.tableWidget.setColumnWidth(2,85)
        self.conexion = conn
        self.nro_orden = orden
        self.manual = None 
        self.nro_reingreso = None 
        self.r_cambio.setChecked(True)
        self.carpeta = None


        self.inicializar()
        self.btn_registrar.clicked.connect(self.registrar)
        self.btn_agregar.clicked.connect(self.agregar)
        self.btn_eliminar.clicked.connect(self.eliminar)
        self.btn_atras.clicked.connect(self.atras)
        self.btn_atras.setIcon(QIcon('icono_imagen/atras.ico'))


    def inicializar(self):
        actual = os.path.abspath(os.getcwd())
        self.carpeta = actual.replace('\\' , '/')

        self.lb_proceso.setText(self.seleccion)

        if self.seleccion == 'DIMENSIONADO':
            try:
                resultado = self.conexion.root.buscar_orden_dim_numero(self.nro_orden)
                if resultado != None :

                    if resultado[18]: #si es manual
                        self.manual = True
                        if resultado[7]:
                            self.lb_doc.setText( resultado[7] ) #TIPO DOCUMENTO
                        else:
                            self.lb_doc.setText( 'No asignado' ) #TIPO DOCUMENTO
                        if resultado[8]:
                            self.lb_documento.setText( str(resultado[8]) ) #nro DOCUMENTO
                        else:
                            self.lb_documento.setText( 'No asignado' ) #nro DOCUMENTO
                    else:

                        self.lb_doc.setText( resultado[7] ) 
                        self.lb_documento.setText( str(resultado[8]) )

                    detalle = json.loads(resultado[6])

                    cantidades = detalle["cantidades"]
                    descripciones = detalle["descripciones"]
                    netos = detalle["valores_neto"]
                    j = 0
                    while j < len( cantidades ):
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( descripciones[j] ) ) 
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem( str( cantidades[j] )  ) )
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( netos[j] )  ) )
                        j+=1      
                                 
            except EOFError:
                QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')

        else:
            
            try:
                if self.seleccion == 'ELABORACION':
                    resultado = self.conexion.root.buscar_orden_elab_numero(self.nro_orden)
                elif self.seleccion == 'CARPINTERIA':
                    resultado = self.conexion.root.buscar_orden_carp_numero(self.nro_orden)
                elif self.seleccion == 'PALLETS':
                    resultado = self.conexion.root.buscar_orden_pall_numero(self.nro_orden)

                if resultado != None :
                    
                    if resultado[15]: #si es manual
                        self.manual = True

                        if resultado[6]:
                            self.lb_doc.setText( resultado[6] ) #TIPO DOCUMENTO
                        else:
                            self.lb_doc.setText( 'No asignado' ) #TIPO DOCUMENTO

                        if resultado[5]:
                            self.lb_documento.setText( str(resultado[5]) ) #nro DOCUMENTO
                        else:
                            self.lb_documento.setText( 'No asignado' ) #nro DOCUMENTO
                    else:                        
                        self.lb_doc.setText( resultado[6] )                 #TIPO DOCUMENTO
                        self.lb_documento.setText( str(resultado[5]) )       #NUMERO DOCUMENTO
                    
                              
                    detalle = json.loads(resultado[11])                  #DETALLE
                    cantidades = detalle["cantidades"]
                    descripciones = detalle["descripciones"]
                    netos = detalle["valores_neto"]
                    j = 0
                    while j < len( cantidades ):
                        fila = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(fila)
                        self.tableWidget.setItem(fila , 0 , QTableWidgetItem( descripciones[j] ) ) 
                        self.tableWidget.setItem(fila , 1 , QTableWidgetItem( str( cantidades[j] )  ) )
                        self.tableWidget.setItem(fila , 2 , QTableWidgetItem( str( netos[j] )  ) )
                        j+=1     

            except EOFError:
                QMessageBox.about(self, 'ERROR', 'Se perdio la conexion con el servidor')

    def agregar(self):
        if self.tableWidget.rowCount() <=16 :
            fila = self.tableWidget.rowCount()
            self.tableWidget.insertRow(fila)
        else:
            QMessageBox.about(self, 'ERROR', 'Ha alcanzado el limite maximo de filas. Intente crear otra Orden para continuar agregando items.')

    def eliminar(self):
        fila = self.tableWidget.currentRow()  #FILA SELECCIONADA , retorna -1 si no se selecciona una fila
        if fila != -1:
            #print('Eliminando la fila ' + str(fila))
            self.tableWidget.removeRow(fila)
        else: 
            QMessageBox.about(self,'Consejo', 'Seleccione una fila para eliminar')
    
    def registrar(self):

        tipo_doc = self.lb_doc.text()
        try:
            nro_doc = int( self.lb_documento.text() )
        except ValueError:
            nro_doc = None
            
        fecha = datetime.now().date()
        motivo = ''
        if self.r_cambio.isChecked():
            motivo = 'CAMBIO'
        elif self.r_devolucion.isChecked():
            motivo = 'DEVOLUCION'
        elif self.r_otro.isChecked():
            motivo = self.txt_otro.text()
        
        proceso = self.seleccion
        descr = self.txt_descripcion.toPlainText()
        solucion = self.txt_solucion.toPlainText()
        lineas = 0

        if motivo != '' and descr != '' and solucion != '' :
            cant = self.tableWidget.rowCount()
            vacias = False #Determinna si existen campos vacios
            correcto = True #Determina si los datos estan escritos correctamente. campos cantidad y valor son numeros.
            cantidades = []
            descripciones = []
            valores_neto = []
            i = 0
            while i< cant:
                descripcion = self.tableWidget.item(i,0) #Collumna descripcion
                cantidad = self.tableWidget.item(i,1) #Columna cantidad
                neto = self.tableWidget.item(i,2) #Columna de valor neto
                if cantidad != None and descripcion != None and neto != None :  #Checkea si se creo una fila, esta no este vacia.
                    if cantidad.text() != '' and descripcion.text() != '' and neto.text() != '' :  #Chekea si se modifico una fila, esta no este vacia
                        try: 
                            nueva_cant = cantidad.text().replace(',','.',3)
                            nuevo_neto = neto.text().replace(',','.',3)
                            cantidades.append( float(nueva_cant) )
                            descripciones.append(descripcion.text())
                            linea = self.separar(descripcion.text(), 60 )
                            lineas += len(linea)
                            print('total de lineas: '+ str(lineas))
                            valores_neto.append(float(nuevo_neto))

                        except ValueError:
                            correcto = False
                    else:
                        vacias=True
                else:
                    vacias = True
                i+=1
            if vacias:
                QMessageBox.about(self, 'Alerta' ,'Una fila y/o columna esta vacia, rellenela para continuar' )
            elif correcto == False:
                QMessageBox.about(self,'Alerta', 'Se encontro un error en una de las cantidades o Valores neto ingresados. Solo ingrese numeros en dichos campos')
            elif lineas > 4:
                QMessageBox.about(self, 'Alerta' ,'El maximo de filas por el formato de impresion es de 4.' )
            else:
                #print(cantidades)
                #print(valores_neto)
                formato = {
                        "cantidades" : cantidades,
                        "descripciones" : descripciones,
                        "valores_neto": valores_neto
                    }
                detalle = json.dumps(formato)
                '''print(fecha)
                print(str(type(fecha)))
                print(str(type(self.nro_orden)))
                print(fecha)
                print(tipo_doc)
                print(nro_doc)
                print(self.nro_orden)
                print(motivo)
                print(descr)
                print(proceso)       
                print(solucion)
                print(str(type(detalle)))'''

                if self.conexion.root.registrar_reingreso( str(fecha), tipo_doc, nro_doc, self.nro_orden, motivo, descr, proceso, detalle,solucion):
                    resultado = self.conexion.root.obtener_max_reingreso()
                    self.nro_reingreso = resultado[0]
                    print('max nro reingreso: ' + str(resultado[0]) + ' de tipo: ' + str(type(resultado[0])))
                    datos = (resultado[0], str(fecha) , tipo_doc , nro_doc , motivo , descr , proceso , solucion, cantidades, descripciones, valores_neto)
                    self.generar_pdf(datos)

                    boton = QMessageBox.question(self, 'Reingreso registrado correctamente', 'Desea ver el reingreso?')
                    if boton == QMessageBox.Yes:
                        self.abrir_pdf()

                    self.hide()
                    self.parent().show()
                else:
                    QMessageBox.about(self,'ERROR','404 NOT FOUND. Contacte con Don Huber ...problemas al registrar')

        else:
            QMessageBox.about(self,'Datos incompletos','Los campos "descripcion" , "solucion" son obligatiorios. Como tambien si selecciona "OTROS" debe rellenar su campo')

    def generar_pdf(self,datos):
        
        documento = canvas.Canvas(self.carpeta +'/reingresos/reingreso_' + str(datos[0]) + '.pdf')
        imagen =  self.carpeta + "/formatos/reingreso_solo.jpg" 

        

        documento.setPageSize(( 216 * mm , 279 * mm))
        documento.drawImage( imagen, 0* mm , 0 * mm , 216 *mm , 139.5 *mm )
        documento.drawImage( imagen, 0* mm , 139.5 * mm , 216 *mm , 139.5 *mm )
        documento.drawString( 0*mm , 139.5 *mm , '------------------------------------------------------------------------------')
        documento.drawString( 108*mm , 139.5 *mm , '----------------------------------------------------------------------------')
        salto = 0
      
        for i in range(2):
            documento.setFont('Helvetica',9)
            
            if datos[2] == 'FACTURA':
                documento.drawString(129*mm, (106.5 + salto )*mm , str(datos[3])  )   #NUMERO DOCUMENTO , FACTURA
            elif datos[2] == 'BOLETA':
                documento.drawString(52* mm, (106.5+ salto )*mm , str(datos[3])  )   #NUMERO DOCUMENTO , BOLETA
           
            lista = self.separar(datos[5],94) #DESCRIPCION
            
            print(len( datos[5] ))
            k = 0 
            j = 0 
            for item in lista:
                documento.drawString(20* mm, (92.5 + salto - k)*mm , lista[j] )  #descripcion del problema
                #documento.drawString(20* mm, (92.5 + salto - k )*mm , descr )  #descripcion del problema
                k += 6
                j += 1
            
            lista = self.separar( datos[7] , 85) #SOLUCION
            print(len(datos[7]))
            k = 0 
            j = 0
            for item in lista:
                documento.drawString(40*mm, (39 + salto - k )*mm , lista[j] )  #solucion del problema
                k += 6
                j += 1
            
            cants = datos[8]
            descrs = datos[9]
            netos = datos[10]
            p = 0
            q = 0
            for item in cants:
                documento.drawString(150*mm, (65 + salto - q )*mm , str(cants[p]) )  #cantidad
                documento.drawString(170*mm, (65 + salto - q )*mm , str(netos[p]) )  #neto
                cadenas = self.separar(descrs[p] , 60 ) 
                for cadena in cadenas:
                    documento.drawString(20*mm, (65 + salto -q )*mm , cadena)  #descripcion
                    q += 5
                p +=1

            documento.setFont('Helvetica-Bold', 9 )

            documento.drawString(177*mm, (124.5 + salto )*mm ,  str(datos[0]) )  #NRO DE REINGRESO
            documento.drawString(177*mm, (115.5 + salto ) *mm , datos[1] )    #FECHA DEL REINGRESO

            if datos[6] == 'DIMENSIONADO':
                documento.drawString(73*mm, (77 + salto )*mm , 'X' )  #PROCESO DIMENSIONADO
            elif datos[6] == 'ELABORACION':    
                documento.drawString(150*mm, (77 + salto )*mm , 'X' )  #PROCESO ELABORACION
            elif datos[6] == 'CARPINTERIA':
                documento.drawString(111*mm, (77 + salto )*mm , 'X' )  #PROCESO CARPINTERIA
            elif datos[6] == 'PALLETS':
                documento.drawString(178.5*mm, (77 + salto )*mm , 'X' )  #PROCESO PALLETS


            if datos[4] == 'CAMBIO':
                documento.drawString(53*mm, (99.5 + salto )*mm , 'X' )  #motivo cambio
            elif datos[4] == 'DEVOLUCION':
                documento.drawString(92*mm, (99.5 + salto )*mm , 'X' )   #motivo devolucion
            else:
                documento.setFont('Helvetica',9)
                documento.drawString(128*mm, (99.5 + salto )*mm , datos[4] )  #motivo otro

            salto += 139.5 
        documento.save()
    
    def abrir_pdf(self):
        ruta = self.carpeta + '/reingresos/reingreso_' + str(self.nro_reingreso) + '.pdf'
        subprocess.Popen([ruta], shell=True)


    def separar(self, cadena, long):
        lista = []
        iter = len(cadena)/long
        iter = int(iter) + 1 #cantidad de items a escribir
        #print('----------------------------------------------------')
        print(cadena)
        print('espacios necesarios: ' + str(iter))
        i = 0
        while len(cadena)> long:
        
            #print('long > '+ str(long) +':')
            aux = cadena[0:long]
            index = aux[::-1].find(' ')
        
            aux = aux[:(long-index)]
            #print('Iteracion: '+str(i)+ ': '+ aux)
            lista.append(aux)
            cadena = cadena[long - index :]
            i += 1

        if len(cadena) > 0 :
            vacias = cadena.count(' ')
            if vacias == len(cadena):
                print('item vacio')
                #print('----------------------------------------------------')
            else:
                #print('fin long < '+ str(long) +':' + cadena)
                lista.append(cadena)
                #print('----------------------------------------------------')

        return lista


    def atras(self):
        self.hide()
        self.parent().show()

class InputDialog(QDialog):
    def __init__(self,label1,label2,title ,parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.txt1 = QLineEdit(self)
        self.txt2 = QLineEdit(self)
        buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self);
       
        layout = QFormLayout(self)
        layout.addRow(label1, self.txt1)
        layout.addRow(label2, self.txt2)
        layout.addWidget(buttonBox)

        buttonBox.accepted.connect(self.accept)
        buttonBox.rejected.connect(self.reject)
    
    def getInputs(self):
        return self.txt1.text(), self.txt2.text()

class InputDialog2(QDialog):
    def __init__(self,label1, titulo ,parent=None):
        super().__init__(parent)
        self.setWindowTitle(titulo)
        self.txt1 = QLineEdit(self)
        self.txt1.setEchoMode(QLineEdit.Password)
        buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self);

        layout = QFormLayout(self)
        layout.addRow(label1, self.txt1)
        layout.addWidget(buttonBox)

        buttonBox.accepted.connect(self.accept)
        buttonBox.rejected.connect(self.reject)
    
            
    def getInputs(self):
        return self.txt1.text()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon('icono_imagen/madenco logo.ico'))
    myappid = 'madenco.personal.area' # arbitrary string
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid) 
    inicio = Dimensionado() 
    inicio.show()
    sys.exit(app.exec_())